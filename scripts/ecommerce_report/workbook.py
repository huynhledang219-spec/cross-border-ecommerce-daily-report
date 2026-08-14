"""Template-preserving Excel report export and inspection."""

from __future__ import annotations

import math
import os
import re
import shutil
import zipfile
from collections.abc import Mapping, Sequence
from copy import copy
from dataclasses import dataclass
from numbers import Real
from pathlib import Path
from tempfile import NamedTemporaryFile
from typing import Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

from .amazon import translate_amazon_title_to_chinese
from .config import RuntimeConfig
from .platforms import is_safe_detail_url


REPORT_HEADERS = [
    "排名",
    "近7天重点选品",
    "来源",
    "品名关键词",
    "中文名称",
    "价格(USD)",
    "商品评分",
    "评论数",
    "GMV",
    "7天GMV",
    "7天销量",
    "关联视频",
    "关联达人",
    "商品详情链接",
    "诊断",
]

_HELPER_HEADERS = tuple(f"趋势日{day}" for day in range(1, 8))
_DEFAULT_CHART_EXTENT = (3_513_455, 1_031_240)
_INVENTORY_SOURCE = "你的库存"
_AMAZON_SOURCE = "Amazon"
_SENSITIVE_REPORT_CONTENT = re.compile(
    r"(?ix)("
    r"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b|"
    r"(?<!\d)1[3-9]\d{9}(?!\d)|"
    r"(?:password|passwd|token|cookie|secret|api[-_ ]?key|authorization)\s*[:=]|"
    r"\bbearer\s+\S+|"
    r"https?://[^/\s@]+@|"
    r"[?&](?:token|session|cookie|secret|api[-_]?key|auth)=[^&\s<]+|"
    r"(?<![a-z])[a-z]:[\\/](?:users|documents and settings)[\\/]|"
    r"/(?:home|users|root)/|"
    r"browser[-_ ]?profile|"
    r"模板行|模板续行"
    r")"
)
_SENSITIVE_REPORT_BYTES = re.compile(
    rb"(?ix)("
    rb"\b[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}\b|"
    rb"(?<!\d)1[3-9]\d{9}(?!\d)|"
    rb"(?:password|passwd|token|cookie|secret|api[-_ ]?key|authorization)\s*[:=]|"
    rb"\bbearer\s+\S+|"
    rb"https?://[^/\s@]+@|"
    rb"[?&](?:token|session|cookie|secret|api[-_]?key|auth)=[^&\s<]+|"
    rb"(?<![a-z])[a-z]:[\\/](?:users|documents[ ]and[ ]settings)[\\/]|"
    rb"/(?:home|users|root)/|"
    rb"browser[-_ ]?profile"
    rb")"
)
_PRIVATE_RESIDUE_BYTES = tuple(
    marker.encode("utf-8") for marker in ("模板行", "模板续行")
)


@dataclass(frozen=True)
class ReportInspection:
    """Stable, read-only summary of workbook properties relevant to the Skill."""

    path: Path
    headers: tuple[str, ...]
    source_order: tuple[str, ...]
    hidden_columns: tuple[str, ...]
    chart_count: int
    chart_extents: tuple[tuple[int, int], ...]
    visible_cells_only: tuple[bool, ...]
    formula_errors: tuple[str, ...]


def _excel_value(value: Any) -> Any:
    if value is pd.NA:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    return value.item() if hasattr(value, "item") else value


def _set_literal(cell, value: Any) -> None:
    """Assign collected text without allowing openpyxl to infer a formula."""

    literal = _excel_value(value)
    cell.value = literal
    if isinstance(literal, str):
        cell.data_type = "s"


def _number(value: Any, default: float = -math.inf) -> float:
    numeric = pd.to_numeric(value, errors="coerce")
    return default if pd.isna(numeric) else float(numeric)


def _trend_values(value: Any) -> list[float] | None:
    if value is None or value is pd.NA:
        return None
    if (
        isinstance(value, Sequence)
        and not isinstance(value, (str, bytes, bytearray))
        and (len(value) == 0 or all(item is None for item in value))
    ):
        return None
    if (
        isinstance(value, (str, bytes, bytearray, Mapping))
        or not isinstance(value, Sequence)
        or len(value) != 7
        or any(
            isinstance(item, bool)
            or not isinstance(item, Real)
            or not math.isfinite(float(item))
            or item < 0
            for item in value
        )
    ):
        raise ValueError("7天GMV趋势必须为空或包含7个有限非负数")
    return [float(item) for item in value]


def _validate_optional_detail_url(value: Any) -> None:
    if value is None or value is pd.NA:
        return
    if isinstance(value, str):
        if value == "":
            return
    else:
        try:
            if bool(pd.isna(value)):
                return
        except (TypeError, ValueError):
            pass
    if not is_safe_detail_url(value):
        raise ValueError("商品详情链接必须是安全的绝对 HTTP(S) URL")


def _formula_inventory(workbook) -> frozenset[tuple[str, str, str]]:
    return frozenset(
        (worksheet.title, cell.coordinate, str(cell.value))
        for worksheet in workbook.worksheets
        for row in worksheet.iter_rows()
        for cell in row
        if cell.data_type == "f"
    )


def _has_content_beyond_helper_columns(worksheet) -> bool:
    """Return true for real cell/drawing content after V, ignoring empty dimensions."""

    if worksheet.max_column > 22:
        for row in worksheet.iter_rows(
            min_row=1,
            max_row=worksheet.max_row,
            min_col=23,
            max_col=worksheet.max_column,
        ):
            for cell in row:
                if (
                    cell.value not in (None, "")
                    or cell.hyperlink is not None
                    or cell.comment is not None
                ):
                    return True
    for drawing in (*worksheet._charts, *worksheet._images):
        marker = getattr(getattr(drawing, "anchor", None), "_from", None)
        if marker is not None and marker.col >= 22:
            return True
    return False


def _sensitive_archive_parts(path: Path) -> tuple[str, ...]:
    flagged: list[str] = []
    with zipfile.ZipFile(path) as archive:
        for member in archive.namelist():
            payload = archive.read(member)
            raw_match = _SENSITIVE_REPORT_BYTES.search(payload) or any(
                marker in payload for marker in _PRIVATE_RESIDUE_BYTES
            )
            text_match = False
            encoding = "utf-8"
            if payload.startswith(
                (b"\xff\xfe\x00\x00", b"\x00\x00\xfe\xff")
            ):
                encoding = "utf-32"
            elif payload.startswith((b"\xff\xfe", b"\xfe\xff")):
                encoding = "utf-16"
            try:
                text = payload.decode(encoding)
            except UnicodeDecodeError:
                text = None
            if text is not None:
                text_match = _SENSITIVE_REPORT_CONTENT.search(text) is not None
            if raw_match or text_match:
                flagged.append(member)
    return tuple(flagged)


def _source_priority(source: str, primary_source: str) -> int:
    priorities = {
        _INVENTORY_SOURCE: 0,
        primary_source: 1,
        _AMAZON_SOURCE: 3,
    }
    return priorities.get(source, 2)


def _prepare_records(
    records: pd.DataFrame, primary_source: str
) -> list[dict[str, Any]]:
    primary_source = str(primary_source).strip()
    if not primary_source or primary_source in {_INVENTORY_SOURCE, _AMAZON_SOURCE}:
        raise ValueError("primary_source must identify a product-intelligence platform")
    prepared = [dict(record) for record in records.to_dict(orient="records")]
    primary_records = [
        record for record in prepared if record.get("source") == primary_source
    ]
    if not primary_records:
        raise ValueError("at least one primary platform row is required")
    if not any(record.get("source") == _AMAZON_SOURCE for record in prepared):
        raise ValueError("at least one Amazon row is required")
    for record in prepared:
        _validate_optional_detail_url(record.get("detail_url"))
    top_records = sorted(
        primary_records,
        key=lambda record: _number(record.get("gmv_7d")),
        reverse=True,
    )[:20]
    for record in top_records:
        _trend_values(record.get("gmv_trend_7d"))
    top_positions = {id(record): position for position, record in enumerate(top_records, 1)}

    for record in prepared:
        position = top_positions.get(id(record))
        record["_top_position"] = position
        record["近7天重点选品"] = f"Top {position}" if position else None
        if record.get("source") == _AMAZON_SOURCE:
            complete_title = " ".join(str(record.get("name") or "").split())
            record["name_cn"] = (
                translate_amazon_title_to_chinese(complete_title)
                if complete_title
                else ""
            )

    def sort_key(record: dict[str, Any]) -> tuple[float, float, float, float]:
        source = str(record.get("source") or "")
        source_priority = float(_source_priority(source, primary_source))
        top_position = record.get("_top_position")
        top_group = 0.0 if top_position else 1.0
        top_order = float(top_position or 999)
        total_gmv = -_number(record.get("gmv"))
        if source != primary_source:
            top_group = 0.0
            top_order = 0.0
        return source_priority, top_group, top_order, total_gmv

    return sorted(prepared, key=sort_key)


def _report_row(record: dict[str, Any]) -> dict[str, Any]:
    return {
        "排名": record.get("rank"),
        "近7天重点选品": record.get("近7天重点选品"),
        "来源": record.get("source"),
        "品名关键词": record.get("name"),
        "中文名称": record.get("name_cn"),
        "价格(USD)": record.get("price"),
        "商品评分": record.get("rating"),
        "评论数": record.get("reviews"),
        "GMV": record.get("gmv"),
        "7天GMV": record.get("gmv_7d"),
        "7天销量": record.get("sold_7d"),
        "关联视频": record.get("videos"),
        "关联达人": record.get("creators"),
        "商品详情链接": record.get("detail_url"),
        "诊断": record.get("diagnostic"),
    }


def _copy_row_style(worksheet, source_row: int, target_row: int) -> None:
    worksheet.row_dimensions[target_row].height = worksheet.row_dimensions[source_row].height
    for column in range(1, len(REPORT_HEADERS) + 1):
        source = worksheet.cell(source_row, column)
        target = worksheet.cell(target_row, column)
        target._style = copy(source._style)
        target.number_format = source.number_format
        target.protection = copy(source.protection)
        target.alignment = copy(source.alignment)


def write_report(
    records: pd.DataFrame,
    output_path: Path,
    template_path: Path,
    *,
    primary_source: str = "EchoTik",
) -> Path:
    """Write normalized records into a copy of an existing Excel template."""

    output_path = Path(output_path)
    template_path = Path(template_path)
    if not template_path.is_file():
        raise FileNotFoundError(f"报表格式模板不存在: {template_path}")
    if output_path.resolve() == template_path.resolve():
        raise ValueError("输出路径不能与模板路径相同")
    template_workbook = load_workbook(template_path, data_only=False)
    try:
        template_formulas = _formula_inventory(template_workbook)
    finally:
        template_workbook.close()
    prepared = _prepare_records(records, primary_source)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with NamedTemporaryFile(
        dir=output_path.parent,
        prefix=f".{output_path.stem}-",
        suffix=".xlsx",
        delete=False,
    ) as temporary_file:
        staged_path = Path(temporary_file.name)
    try:
        shutil.copy2(template_path, staged_path)
        workbook = load_workbook(staged_path)
        try:
            worksheet = workbook.active
            reference_extent = _DEFAULT_CHART_EXTENT
            if worksheet._charts:
                anchor = worksheet._charts[0].anchor
                extent = getattr(anchor, "ext", None)
                if extent is not None:
                    reference_extent = (extent.cx, extent.cy)
            worksheet._charts = []

            if worksheet.max_column > len(REPORT_HEADERS):
                worksheet.delete_cols(
                    len(REPORT_HEADERS) + 1,
                    worksheet.max_column - len(REPORT_HEADERS),
                )
            for column, header in enumerate(REPORT_HEADERS, start=1):
                worksheet.cell(1, column).value = header

            existing_last_row = worksheet.max_row
            output_last_row = len(prepared) + 1
            for row in range(2, max(existing_last_row, output_last_row) + 1):
                for column in range(1, len(REPORT_HEADERS) + 1):
                    cell = worksheet.cell(row, column)
                    cell.value = None
                    cell.hyperlink = None

            for row in range(existing_last_row + 1, output_last_row + 1):
                if existing_last_row >= 4:
                    style_source_row = 3 + ((row - 3) % 2)
                else:
                    style_source_row = 3 if existing_last_row >= 3 else 2
                _copy_row_style(worksheet, style_source_row, row)

            for output_row, record in enumerate(prepared, start=2):
                values = _report_row(record)
                for column, header in enumerate(REPORT_HEADERS, start=1):
                    _set_literal(worksheet.cell(output_row, column), values.get(header))

            detail_column = REPORT_HEADERS.index("商品详情链接") + 1
            diagnostic_column = REPORT_HEADERS.index("诊断") + 1
            detail_letter = worksheet.cell(1, detail_column).column_letter
            worksheet.column_dimensions[detail_letter].hidden = True
            for row in range(2, output_last_row + 1):
                cell = worksheet.cell(row, detail_column)
                detail_url = str(cell.value or "").strip()
                if detail_url:
                    cell.hyperlink = detail_url
                    cell.value = ""
                cell.number_format = ";;;"

            helper_start_column = len(REPORT_HEADERS) + 1
            for offset, header in enumerate(_HELPER_HEADERS):
                helper_column = helper_start_column + offset
                worksheet.cell(1, helper_column).value = header
                helper_letter = worksheet.cell(1, helper_column).column_letter
                worksheet.column_dimensions[helper_letter].hidden = True

            for output_row, record in enumerate(prepared, start=2):
                if record.get("_top_position") is None:
                    continue
                trend = _trend_values(record.get("gmv_trend_7d"))
                if trend is None:
                    worksheet.cell(output_row, diagnostic_column).value = "数据为空"
                    continue
                for offset, value in enumerate(trend):
                    worksheet.cell(output_row, helper_start_column + offset).value = value
                worksheet.cell(output_row, diagnostic_column).value = None

                chart = LineChart()
                chart.width = reference_extent[0] / 360_000
                chart.height = reference_extent[1] / 360_000
                chart.legend = None
                chart.y_axis.delete = True
                chart.x_axis.delete = True
                chart.visible_cells_only = False
                chart.add_data(
                    Reference(
                        worksheet,
                        min_col=helper_start_column,
                        max_col=helper_start_column + 6,
                        min_row=output_row,
                        max_row=output_row,
                    ),
                    from_rows=True,
                    titles_from_data=False,
                )
                chart.series[0].graphicalProperties.line.solidFill = "5B4CF0"
                chart.series[0].graphicalProperties.line.width = 22_000
                diagnostic_letter = worksheet.cell(1, diagnostic_column).column_letter
                worksheet.add_chart(chart, f"{diagnostic_letter}{output_row}")

            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = f"A1:O{output_last_row}"
            workbook.save(staged_path)
        finally:
            workbook.close()

        verification_workbook = load_workbook(staged_path, read_only=False, data_only=False)
        try:
            unexpected_formulas = _formula_inventory(verification_workbook) - template_formulas
            if unexpected_formulas:
                raise ValueError("报表包含模板之外的公式")
        finally:
            verification_workbook.close()
        verify_report(staged_path, template_path)
        os.replace(staged_path, output_path)
    finally:
        staged_path.unlink(missing_ok=True)
    return output_path


def inspect_report(path: Path) -> ReportInspection:
    """Inspect workbook structure without changing the report."""

    path = Path(path)
    workbook = load_workbook(path, data_only=False)
    try:
        worksheet = workbook.active
        headers = tuple(worksheet.cell(1, column).value for column in range(1, 16))

        source_column = REPORT_HEADERS.index("来源") + 1
        source_order: list[str] = []
        for row in range(2, worksheet.max_row + 1):
            source = worksheet.cell(row, source_column).value
            if source and source not in source_order:
                source_order.append(str(source))

        hidden_columns: list[str] = []
        for column in range(1, worksheet.max_column + 1):
            letter = worksheet.cell(1, column).column_letter
            header = worksheet.cell(1, column).value
            if worksheet.column_dimensions[letter].hidden and header:
                hidden_columns.append(str(header))

        formula_errors: list[str] = []
        for audit_sheet in workbook.worksheets:
            for row in audit_sheet.iter_rows():
                for cell in row:
                    if cell.data_type in {"e", "f"}:
                        formula_errors.append(
                            f"{audit_sheet.title}!{cell.coordinate}:{cell.value}"
                        )

        charts = tuple(worksheet._charts)
        return ReportInspection(
            path=path,
            headers=tuple(str(header) for header in headers),
            source_order=tuple(source_order),
            hidden_columns=tuple(hidden_columns),
            chart_count=len(charts),
            chart_extents=tuple(
                (chart.anchor.ext.cx, chart.anchor.ext.cy) for chart in charts
            ),
            visible_cells_only=tuple(chart.visible_cells_only for chart in charts),
            formula_errors=tuple(formula_errors),
        )
    finally:
        workbook.close()


def verify_report(path: Path, template_path: Path | None = None) -> ReportInspection:
    """Fail closed when a report does not satisfy the public workbook contract."""

    verified_path = RuntimeConfig.ensure_outside_skill(path, "report output")
    reference_template = Path(template_path) if template_path is not None else (
        RuntimeConfig._SKILL_DIRECTORY / "assets" / "report-template.xlsx"
    )
    template_workbook = load_workbook(reference_template, data_only=False)
    try:
        if not template_workbook.active._charts:
            raise ValueError("模板缺少参考图表")
        template_anchor = template_workbook.active._charts[0].anchor
        template_extent = (template_anchor.ext.cx, template_anchor.ext.cy)
    finally:
        template_workbook.close()
    inspection = inspect_report(verified_path)
    if inspection.headers != tuple(REPORT_HEADERS):
        raise ValueError("报表表头不符合公开格式")
    required_hidden = ("商品详情链接",) + _HELPER_HEADERS
    if inspection.hidden_columns != required_hidden:
        raise ValueError("报表隐藏列不符合公开格式")
    if inspection.formula_errors:
        raise ValueError("报表包含公式或公式错误")
    if any(extent != template_extent for extent in inspection.chart_extents):
        raise ValueError("报表图表尺寸与模板图表尺寸不一致")
    workbook = load_workbook(verified_path, data_only=False)
    try:
        worksheet = workbook.active
        if _has_content_beyond_helper_columns(worksheet):
            raise ValueError("报表在V列之后包含额外内容或图形")
        detail_column = REPORT_HEADERS.index("商品详情链接") + 1
        for row in range(2, worksheet.max_row + 1):
            detail_cell = worksheet.cell(row, detail_column)
            detail_value = detail_cell.value
            if detail_value not in (None, ""):
                if not is_safe_detail_url(detail_value) or detail_cell.hyperlink is None:
                    raise ValueError("报表商品详情链接目标不安全")
                detail_target = detail_cell.hyperlink.target
                if (
                    not is_safe_detail_url(detail_target)
                    or detail_target != detail_value
                ):
                    raise ValueError("报表商品详情链接目标不安全")
            elif detail_cell.hyperlink is not None:
                raise ValueError("报表商品详情链接目标不安全")
        populated_rows = [
            row
            for row in range(2, worksheet.max_row + 1)
            if any(
                worksheet.cell(row, column).value not in (None, "")
                for column in range(1, len(REPORT_HEADERS) + 1)
            )
        ]
        expected_filter = f"A1:O{max(populated_rows, default=1)}"
        if str(worksheet.freeze_panes or "") != "A2" or worksheet.auto_filter.ref != expected_filter:
            raise ValueError("报表冻结窗格或筛选范围无效")
        source_values = [
            str(worksheet.cell(row, 3).value)
            for row in populated_rows
            if worksheet.cell(row, 3).value not in (None, "")
        ]
        top_rows = [
            row
            for row in range(2, worksheet.max_row + 1)
            if worksheet.cell(row, 2).value not in (None, "")
        ]
        if not top_rows:
            raise ValueError("at least one primary platform row is required")
        if _AMAZON_SOURCE not in source_values:
            raise ValueError("at least one Amazon row is required")
        non_reserved_sources = {
            source
            for source in source_values
            if source not in {_INVENTORY_SOURCE, _AMAZON_SOURCE}
        }
        if top_rows:
            top_sources = {
                str(worksheet.cell(row, 3).value or "").strip() for row in top_rows
            }
            reserved_sources = {_INVENTORY_SOURCE, _AMAZON_SOURCE, ""}
            if len(top_sources) != 1 or top_sources & reserved_sources:
                raise ValueError("invalid primary platform source in Top rows")
            primary_source = next(iter(top_sources))
            if non_reserved_sources != {primary_source}:
                raise ValueError("invalid primary platform source outside Top rows")
        else:
            if non_reserved_sources:
                raise ValueError("primary platform source requires Top rows")
            primary_source = ""
        allowed_sources = {_INVENTORY_SOURCE, _AMAZON_SOURCE}
        if primary_source:
            allowed_sources.add(primary_source)
        if any(source not in allowed_sources for source in inspection.source_order):
            raise ValueError("报表来源顺序无效")
        source_priorities = [
            _source_priority(source, primary_source) for source in source_values
        ]
        if source_priorities != sorted(source_priorities):
            raise ValueError("报表来源连续分组顺序无效")
        if list(inspection.source_order) != sorted(
            inspection.source_order,
            key=lambda source: _source_priority(source, primary_source),
        ):
            raise ValueError("报表来源顺序无效")
        top_labels = [worksheet.cell(row, 2).value for row in top_rows]
        expected_labels = [f"Top {position}" for position in range(1, len(top_rows) + 1)]
        top_gmvs = [_number(worksheet.cell(row, 10).value) for row in top_rows]
        if (
            len(top_rows) > 20
            or top_labels != expected_labels
            or any(worksheet.cell(row, 3).value != primary_source for row in top_rows)
            or any(value == -math.inf for value in top_gmvs)
            or top_gmvs != sorted(top_gmvs, reverse=True)
        ):
            raise ValueError("Primary platform Top 20 标签或7天GMV顺序无效")

        top_trends = {
            row: _trend_values(
                [worksheet.cell(row, column).value for column in range(16, 23)]
            )
            for row in top_rows
        }
        chart_rows: list[int] = []
        for chart in worksheet._charts:
            chart_row = chart.anchor._from.row + 1
            chart_rows.append(chart_row)
            expected_range = f"${worksheet.cell(1, 16).column_letter}${chart_row}:${worksheet.cell(1, 22).column_letter}${chart_row}"
            series_range = (
                chart.series[0].val.numRef.f if len(chart.series) == 1 else ""
            )
            trend = top_trends.get(chart_row)
            if (
                chart_row not in top_rows
                or chart.anchor._from.col != 14
                or expected_range not in str(series_range)
                or trend is None
                or chart.visible_cells_only is not False
            ):
                raise ValueError("Top 20 趋势图与数据行不一致")
        if len(chart_rows) > 20 or len(chart_rows) != len(set(chart_rows)):
            raise ValueError("Top 20 趋势图数量或对应行无效")
        if inspection.chart_extents and len(set(inspection.chart_extents)) != 1:
            raise ValueError("Top 20 趋势图尺寸不一致")
        for row in top_rows:
            diagnostic = worksheet.cell(row, 15).value
            has_chart = row in chart_rows
            has_trend = top_trends[row] is not None
            if (
                has_trend != has_chart
                or (not has_trend and diagnostic != "数据为空")
                or (has_trend and diagnostic == "数据为空")
            ):
                raise ValueError("Top 20 行必须具有趋势图或数据为空诊断")
        for row in range(2, worksheet.max_row + 1):
            if worksheet.cell(row, 3).value != _AMAZON_SOURCE:
                continue
            original_title = str(worksheet.cell(row, 4).value or "").strip()
            chinese_title = str(worksheet.cell(row, 5).value or "").strip()
            if (
                not original_title
                or not chinese_title
                or not re.search(r"[\u3400-\u4dbf\u4e00-\u9fff]", chinese_title)
                or chinese_title.casefold() == original_title.casefold()
                or original_title.endswith(("...", "…"))
                or chinese_title.endswith(("...", "…"))
            ):
                raise ValueError("Amazon 中文全称或英文全称缺失、截断")
    finally:
        workbook.close()
    if _sensitive_archive_parts(verified_path):
        raise ValueError("报表包含敏感账户或本地运行内容")
    return inspection


__all__ = [
    "REPORT_HEADERS",
    "ReportInspection",
    "inspect_report",
    "verify_report",
    "write_report",
]
