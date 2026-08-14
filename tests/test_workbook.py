from __future__ import annotations

import unittest
import zipfile
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import patch

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, PatternFill
from openpyxl.worksheet.hyperlink import Hyperlink

from scripts.ecommerce_report import workbook as workbook_module
from scripts.ecommerce_report.workbook import REPORT_HEADERS, inspect_report, write_report


REFERENCE_CHART_EXTENT = (4_200_000, 1_260_000)


def create_synthetic_template(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "每日选品"
    template_headers = REPORT_HEADERS + ["7天GMV较昨日", "7天销量较昨日"]
    worksheet.append(template_headers)
    worksheet.append(["模板行"] * len(template_headers))
    worksheet.append(["模板续行"] * len(template_headers))
    worksheet.append(["模板交替行"] * len(template_headers))

    gold_fill = PatternFill(fill_type="solid", fgColor="FFD700")
    for cell in worksheet[1]:
        cell.fill = gold_fill
    for row in (2, 3, 4):
        for column in range(1, len(template_headers) + 1):
            worksheet.cell(row, column).alignment = Alignment(horizontal="left")
    gray_fill = PatternFill(fill_type="solid", fgColor="FFF2F2F2")
    for cell in worksheet[4]:
        cell.fill = gray_fill
    worksheet.column_dimensions["D"].width = 31.5
    worksheet.column_dimensions["O"].width = 24.0
    worksheet.row_dimensions[2].height = 36.0
    worksheet.row_dimensions[3].height = 29.0
    worksheet.row_dimensions[4].height = 29.0

    for column, value in enumerate(range(1, 8), start=16):
        worksheet.cell(2, column, value)
    chart = LineChart()
    chart.width = REFERENCE_CHART_EXTENT[0] / 360_000
    chart.height = REFERENCE_CHART_EXTENT[1] / 360_000
    chart.add_data(
        Reference(worksheet, min_col=16, max_col=22, min_row=2, max_row=2),
        from_rows=True,
    )
    worksheet.add_chart(chart, "O2")
    workbook.save(path)
    workbook.close()


def normalized_records(primary_source: str = "EchoTik") -> pd.DataFrame:
    records: list[dict] = [
        {
            "rank": "SKU",
            "source": "你的库存",
            "name": "Inventory SKU",
            "name_cn": "库存商品",
            "price": 22.99,
            "diagnostic": "库存诊断",
        }
    ]
    records.extend(
        {
            "rank": index,
            "source": primary_source,
            "name": f"Echo product {index}",
            "name_cn": f"页面中文名 {index}",
            "price": 10 + index,
            "rating": 4.5,
            "reviews": index,
            "gmv": 10_000 - index,
            "gmv_7d": index,
            "sold_7d": index * 2,
            "videos": index + 3,
            "creators": index + 4,
            "detail_url": f"https://echotik.example/product/{index}",
            "diagnostic": f"诊断 {index}",
            "gmv_trend_7d": [1, 2, 3, 4, 5, 6, index] if index != 22 else None,
        }
        for index in range(1, 23)
    )
    records.append(
        {
            "rank": 23,
            "source": "Amazon",
            "name": "Complete English Product Title With Every Important Detail",
            "name_cn": "不得采用的已有短名",
            "price": 29.99,
            "rating": 4.8,
            "reviews": 1200,
        }
    )
    return pd.DataFrame(records)


class WorkbookExportTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = TemporaryDirectory()
        self.root = Path(self.temporary_directory.name)
        self.template_path = self.root / "synthetic-template.xlsx"
        self.output_path = self.root / "output" / "daily-report.xlsx"
        create_synthetic_template(self.template_path)

    def tearDown(self) -> None:
        self.temporary_directory.cleanup()

    def write_fixture(self) -> Path:
        with patch(
            "scripts.ecommerce_report.workbook.translate_amazon_title_to_chinese",
            return_value="包含每个重要细节的完整英文商品标题",
        ) as translate:
            result = write_report(
                normalized_records(), self.output_path, self.template_path
            )
        translate.assert_called_once_with(
            "Complete English Product Title With Every Important Detail"
        )
        return result

    def test_orders_sources_and_places_primary_top_twenty_by_seven_day_gmv(self) -> None:
        result = self.write_fixture()
        worksheet = load_workbook(result, data_only=False).active

        sources = [worksheet.cell(row, 3).value for row in range(2, 26)]
        self.assertEqual(sources[0], "你的库存")
        self.assertEqual(sources[1:23], ["EchoTik"] * 22)
        self.assertEqual(sources[23], "Amazon")
        self.assertEqual(
            [worksheet.cell(row, 10).value for row in range(3, 23)],
            list(range(22, 2, -1)),
        )
        self.assertEqual(
            [worksheet.cell(row, 2).value for row in range(3, 23)],
            [f"Top {position}" for position in range(1, 21)],
        )
        self.assertIsNone(worksheet.cell(23, 2).value)
        self.assertIsNone(worksheet.cell(24, 2).value)
        worksheet.parent.close()

    def test_preserves_complete_names_and_removes_deleted_comparison_columns(self) -> None:
        result = self.write_fixture()
        worksheet = load_workbook(result, data_only=False).active
        headers = [worksheet.cell(1, column).value for column in range(1, 16)]

        self.assertEqual(headers, REPORT_HEADERS)
        self.assertNotIn("7天GMV较昨日", headers)
        self.assertNotIn("7天销量较昨日", headers)
        self.assertEqual(worksheet["E3"].value, "页面中文名 22")
        self.assertEqual(worksheet["E25"].value, "包含每个重要细节的完整英文商品标题")

        detail_column = worksheet.column_dimensions["N"]
        self.assertTrue(detail_column.hidden)
        self.assertEqual(
            worksheet["N3"].hyperlink.target,
            "https://echotik.example/product/22",
        )
        self.assertTrue(
            all(
                worksheet.column_dimensions[worksheet.cell(1, column).column_letter].hidden
                for column in range(16, 23)
            )
        )
        worksheet.parent.close()

    def test_preserves_template_dimensions_alignment_and_chart_contract(self) -> None:
        result = self.write_fixture()
        worksheet = load_workbook(result, data_only=False).active

        self.assertEqual(worksheet["A1"].fill.fgColor.rgb[-6:], "FFD700")
        self.assertEqual(worksheet.column_dimensions["D"].width, 31.5)
        self.assertEqual(worksheet.column_dimensions["O"].width, 24.0)
        self.assertEqual(worksheet.row_dimensions[2].height, 36.0)
        self.assertEqual(worksheet.row_dimensions[25].height, 29.0)
        self.assertEqual(worksheet["A2"].alignment.horizontal, "left")
        self.assertTrue(
            all(
                worksheet.cell(row, column).alignment.horizontal == "left"
                for row in range(2, 26)
                for column in range(1, 16)
            )
        )
        self.assertEqual(len(worksheet._charts), 19)
        self.assertTrue(
            all(
                (chart.anchor.ext.cx, chart.anchor.ext.cy) == REFERENCE_CHART_EXTENT
                for chart in worksheet._charts
            )
        )
        self.assertTrue(all(chart.visible_cells_only is False for chart in worksheet._charts))
        self.assertTrue(all(chart.anchor._from.col == 14 for chart in worksheet._charts))
        self.assertEqual(
            [chart.anchor._from.row + 1 for chart in worksheet._charts],
            list(range(4, 23)),
        )
        self.assertEqual(
            [chart.series[0].val.numRef.f for chart in worksheet._charts],
            [f"'每日选品'!$P${row}:$V${row}" for row in range(4, 23)],
        )
        worksheet.parent.close()

    def test_repeats_the_two_actual_body_row_styles_in_order(self) -> None:
        result = self.write_fixture()
        worksheet = load_workbook(result, data_only=False).active

        self.assertNotEqual(worksheet["A3"].fill.fgColor.rgb, worksheet["A4"].fill.fgColor.rgb)
        self.assertEqual(worksheet["A5"].fill.fgColor.rgb, worksheet["A3"].fill.fgColor.rgb)
        self.assertEqual(worksheet["A6"].fill.fgColor.rgb, worksheet["A4"].fill.fgColor.rgb)
        worksheet.parent.close()

    def test_missing_trend_writes_data_empty_without_a_chart(self) -> None:
        result = self.write_fixture()
        worksheet = load_workbook(result, data_only=False).active

        self.assertEqual(worksheet["B3"].value, "Top 1")
        self.assertEqual(worksheet["O3"].value, "数据为空")
        chart_rows = {chart.anchor._from.row + 1 for chart in worksheet._charts}
        self.assertNotIn(3, chart_rows)
        worksheet.parent.close()

    def test_empty_trend_writes_data_empty_without_a_chart(self) -> None:
        records = normalized_records()
        records.at[22, "gmv_trend_7d"] = []
        with patch(
            "scripts.ecommerce_report.workbook.translate_amazon_title_to_chinese",
            return_value="包含每个重要细节的完整英文商品标题",
        ):
            result = write_report(records, self.output_path, self.template_path)

        workbook = load_workbook(result, data_only=False)
        try:
            worksheet = workbook.active
            self.assertEqual(worksheet["B3"].value, "Top 1")
            self.assertEqual(worksheet["O3"].value, "数据为空")
            self.assertNotIn(
                3, {chart.anchor._from.row + 1 for chart in worksheet._charts}
            )
        finally:
            workbook.close()

    def test_malformed_present_trend_does_not_replace_an_existing_report(self) -> None:
        invalid_trends = (
            "[1, 2, 3, 4, 5, 6, 7]",
            True,
            -1,
            float("nan"),
            float("inf"),
            [1, 2, 3, 4, 5, 6],
            [1, 2, 3, 4, 5, 6, True],
            [1, 2, 3, 4, 5, 6, "7"],
            [1, 2, 3, 4, 5, 6, float("nan")],
            [1, 2, 3, 4, 5, 6, float("inf")],
            [1, 2, 3, 4, 5, 6, -1],
        )
        for invalid_trend in invalid_trends:
            with self.subTest(trend=invalid_trend):
                self.output_path.parent.mkdir(parents=True, exist_ok=True)
                original = b"existing verified report"
                self.output_path.write_bytes(original)
                records = normalized_records()
                records.at[22, "gmv_trend_7d"] = invalid_trend

                with patch(
                    "scripts.ecommerce_report.workbook.translate_amazon_title_to_chinese",
                    return_value="包含每个重要细节的完整英文商品标题",
                ):
                    with self.assertRaisesRegex(ValueError, "7天GMV趋势"):
                        write_report(records, self.output_path, self.template_path)
                self.assertEqual(self.output_path.read_bytes(), original)

    def test_verifier_never_certifies_partial_trend_data_as_empty(self) -> None:
        result = self.write_fixture()
        workbook = load_workbook(result)
        workbook.active["P3"] = 1
        workbook.save(result)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "7天GMV趋势"):
            workbook_module.verify_report(result, self.template_path)

    def test_inspect_report_returns_a_stable_public_audit_summary(self) -> None:
        result = self.write_fixture()

        inspection = inspect_report(result)

        self.assertEqual(inspection.path, result)
        self.assertEqual(inspection.headers, tuple(REPORT_HEADERS))
        self.assertEqual(inspection.source_order, ("你的库存", "EchoTik", "Amazon"))
        self.assertEqual(
            inspection.hidden_columns,
            ("商品详情链接", "趋势日1", "趋势日2", "趋势日3", "趋势日4", "趋势日5", "趋势日6", "趋势日7"),
        )
        self.assertEqual(inspection.chart_count, 19)
        self.assertEqual(
            inspection.chart_extents,
            (REFERENCE_CHART_EXTENT,) * 19,
        )
        self.assertEqual(inspection.visible_cells_only, (False,) * 19)
        self.assertEqual(inspection.formula_errors, ())

    def test_verify_report_accepts_a_complete_contract_conforming_report(self) -> None:
        """A verifier that cannot certify the generated contract is not usable as a gate."""
        result = self.write_fixture()

        try:
            inspection = workbook_module.verify_report(result, self.template_path)
        except AttributeError:
            self.fail("verify_report is not implemented")

        self.assertEqual(inspection.path, result)

    def test_report_ranks_configured_primary_source_and_keeps_amazon_last(self) -> None:
        records = normalized_records(primary_source="MarketPulse")
        with patch(
            "scripts.ecommerce_report.workbook.translate_amazon_title_to_chinese",
            return_value="包含每个重要细节的完整英文商品标题",
        ):
            result = write_report(
                records,
                self.output_path,
                self.template_path,
                primary_source="MarketPulse",
            )

        workbook = load_workbook(result, data_only=False)
        try:
            worksheet = workbook.active
            self.assertEqual(
                [worksheet.cell(row, 2).value for row in range(3, 23)],
                [f"Top {position}" for position in range(1, 21)],
            )
            self.assertTrue(
                all(
                    worksheet.cell(row, 3).value == "MarketPulse"
                    for row in range(3, 25)
                )
            )
            self.assertEqual(worksheet.cell(25, 3).value, "Amazon")
            self.assertEqual(worksheet.cell(1, 14).value, "商品详情链接")
            self.assertTrue(worksheet.column_dimensions["N"].hidden)
        finally:
            workbook.close()
        self.assertEqual(
            workbook_module.verify_report(result, self.template_path).source_order,
            ("你的库存", "MarketPulse", "Amazon"),
        )

    def test_verifier_rejects_top_rows_from_two_primary_sources(self) -> None:
        records = normalized_records(primary_source="MarketPulse")
        with patch(
            "scripts.ecommerce_report.workbook.translate_amazon_title_to_chinese",
            return_value="包含每个重要细节的完整英文商品标题",
        ):
            result = write_report(
                records,
                self.output_path,
                self.template_path,
                primary_source="MarketPulse",
            )
        workbook = load_workbook(result)
        workbook.active["C4"] = "EchoTik"
        workbook.save(result)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "primary platform source"):
            workbook_module.verify_report(result, self.template_path)

    def test_verify_report_rejects_a_changed_public_header(self) -> None:
        """Relying on a readable XLSX alone would accept a broken column contract."""
        result = self.write_fixture()
        workbook = load_workbook(result)
        workbook.active["J1"] = "错误7天GMV"
        workbook.save(result)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "表头"):
            workbook_module.verify_report(result, self.template_path)

    def test_writer_and_verifier_require_primary_and_amazon_sources(self) -> None:
        self.output_path.parent.mkdir(parents=True, exist_ok=True)
        original = b"existing verified report"
        cases = (
            ("empty", pd.DataFrame(), "primary platform"),
            (
                "missing Amazon",
                normalized_records()
                .query("source != 'Amazon'")
                .reset_index(drop=True),
                "Amazon",
            ),
            (
                "missing primary",
                normalized_records()
                .query("source != 'EchoTik'")
                .reset_index(drop=True),
                "primary platform",
            ),
        )
        for name, records, message in cases:
            with self.subTest(case=name):
                self.output_path.write_bytes(original)
                with self.assertRaisesRegex(ValueError, message):
                    write_report(records, self.output_path, self.template_path)
                self.assertEqual(self.output_path.read_bytes(), original)

        for missing_source in ("Amazon", "EchoTik"):
            with self.subTest(verifier_missing=missing_source):
                result = self.write_fixture()
                workbook = load_workbook(result)
                worksheet = workbook.active
                if missing_source == "Amazon":
                    worksheet.delete_rows(25)
                    worksheet.auto_filter.ref = "A1:O24"
                else:
                    worksheet.delete_rows(3, 22)
                    worksheet._charts = []
                    worksheet.auto_filter.ref = "A1:O3"
                workbook.save(result)
                workbook.close()

                expected_message = (
                    "Amazon" if missing_source == "Amazon" else "primary platform"
                )
                with self.assertRaisesRegex(ValueError, expected_message):
                    workbook_module.verify_report(result, self.template_path)

    def test_verify_report_rejects_wrong_freeze_pane_or_public_filter_range(self) -> None:
        """A valid-looking table must still preserve the promised navigation layout."""
        for mutation in ("freeze", "filter"):
            with self.subTest(mutation=mutation):
                result = self.write_fixture()
                workbook = load_workbook(result)
                worksheet = workbook.active
                if mutation == "freeze":
                    worksheet.freeze_panes = None
                else:
                    worksheet.auto_filter.ref = "A1:N25"
                workbook.save(result)
                workbook.close()

                with self.assertRaisesRegex(ValueError, "冻结窗格或筛选范围"):
                    workbook_module.verify_report(result, self.template_path)

    def test_verify_report_rejects_top_labels_not_matching_descending_gmv(self) -> None:
        """Checking only chart count would accept mislabeled or unsorted Top products."""
        result = self.write_fixture()
        workbook = load_workbook(result)
        worksheet = workbook.active
        worksheet["J3"], worksheet["J4"] = worksheet["J4"].value, worksheet["J3"].value
        workbook.save(result)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "Top 20"):
            workbook_module.verify_report(result, self.template_path)

    def test_verify_report_rejects_a_source_that_reappears_after_a_later_group(self) -> None:
        """Unique first-seen sources cannot detect echotik→Amazon→echotik interleaving."""
        result = self.write_fixture()
        workbook = load_workbook(result)
        workbook.active["C23"] = "Amazon"
        workbook.save(result)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "来源连续分组"):
            workbook_module.verify_report(result, self.template_path)

    def test_verify_report_rejects_a_top_row_without_chart_or_data_empty_diagnostic(self) -> None:
        """Counting charts globally would miss a chart attached to the wrong Top identity."""
        result = self.write_fixture()
        workbook = load_workbook(result)
        worksheet = workbook.active
        worksheet._charts = worksheet._charts[1:]
        workbook.save(result)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "趋势图"):
            workbook_module.verify_report(result, self.template_path)

    def test_verify_report_rejects_uniform_chart_sizes_that_do_not_match_template(self) -> None:
        """Checking only mutual consistency would accept charts resized away from the template."""
        result = self.write_fixture()
        workbook = load_workbook(result)
        for chart in workbook.active._charts:
            chart.anchor.ext.cx = 720_000
            chart.anchor.ext.cy = 360_000
        workbook.save(result)
        workbook.close()

        try:
            workbook_module.verify_report(result, self.template_path)
        except Exception as error:
            self.assertIs(type(error), ValueError)
            self.assertIn("模板图表尺寸", str(error))
        else:
            self.fail("uniform non-template chart sizes were accepted")

    def test_verify_report_rejects_visible_helper_or_detail_columns(self) -> None:
        """Visible helper data would violate the published workbook layout."""
        result = self.write_fixture()
        workbook = load_workbook(result)
        workbook.active.column_dimensions["P"].hidden = False
        workbook.save(result)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "隐藏列"):
            workbook_module.verify_report(result, self.template_path)

    def test_verify_report_rejects_visible_content_beyond_the_helper_columns(
        self,
    ) -> None:
        """A populated W column must not escape the A:V workbook contract."""
        result = self.write_fixture()
        workbook = load_workbook(result)
        workbook.active["W1"] = "unexpected visible report column"
        workbook.save(result)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "V列之后"):
            workbook_module.verify_report(result, self.template_path)

    def test_detail_urls_are_safe_before_write_and_during_verification(self) -> None:
        invalid_urls = (
            "products/42",
            "/products/42",
            "file:///etc/passwd",
            r"\\server\share\product.html",
            "javascript:alert(1)",
            "data:text/html,malicious",
            "https:///products/42",
            "https://user:password@example.test/products/42",
            "https://example.test/path with space",
            r"https://example.test\malformed",
            "https://example.test/products/42\x7fhidden",
            "https://example.test/products/42\u200bhidden",
            "https://example.test/products/42%00hidden",
            "https://example.test/products/42%0d%0ahidden",
            "https://example.test/products/42%E2%80%8Bhidden",
        )
        for invalid_url in invalid_urls:
            with self.subTest(writer_url=invalid_url):
                self.output_path.parent.mkdir(parents=True, exist_ok=True)
                original = b"existing verified report"
                self.output_path.write_bytes(original)
                records = normalized_records()
                records.at[22, "detail_url"] = invalid_url
                with self.assertRaisesRegex(ValueError, "安全的绝对 HTTP"):
                    write_report(records, self.output_path, self.template_path)
                self.assertEqual(self.output_path.read_bytes(), original)

            with self.subTest(verifier_url=invalid_url):
                result = self.write_fixture()
                workbook = load_workbook(result)
                worksheet = workbook.active
                worksheet["N3"].hyperlink = invalid_url
                worksheet["N3"].value = ""
                workbook.save(result)
                workbook.close()
                with self.assertRaisesRegex(ValueError, "详情链接"):
                    workbook_module.verify_report(result, self.template_path)

        result = self.write_fixture()
        workbook = load_workbook(result)
        worksheet = workbook.active
        worksheet["N3"].hyperlink = None
        worksheet["N3"].value = "javascript:alert(1)"
        workbook.save(result)
        workbook.close()
        with self.assertRaisesRegex(ValueError, "详情链接"):
            workbook_module.verify_report(result, self.template_path)

    def test_verifier_requires_matching_external_detail_hyperlinks(self) -> None:
        for case in ("internal", "missing", "mismatch"):
            with self.subTest(case=case):
                result = self.write_fixture()
                workbook = load_workbook(result)
                worksheet = workbook.active
                if case == "internal":
                    worksheet["N3"].hyperlink = Hyperlink(
                        ref="N3", location="Sheet1!A1"
                    )
                elif case == "missing":
                    worksheet["N3"].hyperlink = Hyperlink(ref="N3")
                else:
                    worksheet["N3"].value = "https://example.test/products/42"
                    worksheet["N3"].hyperlink = (
                        "https://example.test/products/different"
                    )
                workbook.save(result)
                workbook.close()

                with self.assertRaisesRegex(ValueError, "详情链接"):
                    workbook_module.verify_report(result, self.template_path)

    def test_verify_report_rejects_a_drawing_anchored_beyond_helper_columns(
        self,
    ) -> None:
        result = self.write_fixture()
        workbook = load_workbook(result)
        worksheet = workbook.active
        unexpected_chart = LineChart()
        unexpected_chart.width = REFERENCE_CHART_EXTENT[0] / 360_000
        unexpected_chart.height = REFERENCE_CHART_EXTENT[1] / 360_000
        worksheet.add_chart(unexpected_chart, "W2")
        workbook.save(result)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "V列之后"):
            workbook_module.verify_report(result, self.template_path)

    def test_verify_report_rejects_missing_or_visibly_truncated_amazon_translation(self) -> None:
        """A source row existing does not prove its promised complete Chinese title exists."""
        for invalid_translation in (None, "被截断的中文名称..."):
            with self.subTest(invalid_translation=invalid_translation):
                result = self.write_fixture()
                workbook = load_workbook(result)
                workbook.active["E25"] = invalid_translation
                workbook.save(result)
                workbook.close()

                with self.assertRaisesRegex(ValueError, "Amazon 中文全称"):
                    workbook_module.verify_report(result, self.template_path)

    def test_verify_report_rejects_a_pure_english_amazon_chinese_name(self) -> None:
        """A copied English title is nonempty but is not the promised Chinese translation."""
        result = self.write_fixture()
        workbook = load_workbook(result)
        worksheet = workbook.active
        worksheet["E25"] = worksheet["D25"].value
        workbook.save(result)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "Amazon 中文全称"):
            workbook_module.verify_report(result, self.template_path)

    def test_verify_report_rejects_an_output_inside_the_skill(self) -> None:
        """A structurally valid workbook is still unsafe when published in the Skill tree."""
        packaged_template = Path(__file__).resolve().parents[1] / "assets" / "report-template.xlsx"

        with self.assertRaisesRegex(ValueError, "Skill directory"):
            workbook_module.verify_report(packaged_template)

    def test_verify_report_rejects_sensitive_account_content(self) -> None:
        """A valid layout must not certify account identifiers embedded in workbook text."""
        result = self.write_fixture()
        workbook = load_workbook(result)
        workbook.active["D25"] = "Contact user@example.com for private access"
        workbook.save(result)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "敏感"):
            workbook_module.verify_report(result, self.template_path)

    def test_verify_report_scans_text_and_raw_bytes_in_every_zip_part(self) -> None:
        """Restricting scans to XML/RELS would miss private residue in custom or binary parts."""
        cases = (
            ("custom/private.txt", b"private contact: hidden@example.invalid", True),
            (
                "custom/private-utf16.txt",
                "private contact: hidden@example.invalid".encode("utf-16"),
                True,
            ),
            ("xl/media/private.bin", b"\x89BIN\xfftoken=private-value\x00", True),
            ("xl/media/safe.bin", b"\x89PNG\r\n\x1a\n\x00\xff\x10\x80", False),
        )
        for member, payload, should_reject in cases:
            with self.subTest(member=member):
                result = self.write_fixture()
                with zipfile.ZipFile(result, "a") as archive:
                    archive.writestr(member, payload)

                if should_reject:
                    with self.assertRaisesRegex(ValueError, "敏感"):
                        workbook_module.verify_report(result, self.template_path)
                else:
                    workbook_module.verify_report(result, self.template_path)

    def test_write_report_runs_verification_before_replacing_existing_output(self) -> None:
        """Leaving verification as an optional command would allow unsafe output publication."""
        self.output_path.parent.mkdir(parents=True)
        original = b"existing verified report"
        self.output_path.write_bytes(original)
        records = normalized_records()
        records.at[22, "name"] = "Contact user@example.com"

        with patch(
            "scripts.ecommerce_report.workbook.translate_amazon_title_to_chinese",
            return_value="包含每个重要细节的完整英文商品标题",
        ):
            with self.assertRaisesRegex(ValueError, "敏感"):
                write_report(records, self.output_path, self.template_path)

        self.assertEqual(self.output_path.read_bytes(), original)

    def test_rejects_using_the_template_as_the_output_without_changing_it(self) -> None:
        original = self.template_path.read_bytes()

        with self.assertRaisesRegex(ValueError, "不能与模板路径相同"):
            write_report(normalized_records(), self.template_path, self.template_path)

        self.assertEqual(self.template_path.read_bytes(), original)

    def test_translation_failure_does_not_replace_an_existing_report(self) -> None:
        self.output_path.parent.mkdir(parents=True)
        original = b"existing valid report"
        self.output_path.write_bytes(original)

        with patch(
            "scripts.ecommerce_report.workbook.translate_amazon_title_to_chinese",
            side_effect=RuntimeError("translation unavailable"),
        ):
            with self.assertRaisesRegex(RuntimeError, "translation unavailable"):
                write_report(normalized_records(), self.output_path, self.template_path)

        self.assertEqual(self.output_path.read_bytes(), original)

    def test_dynamic_web_values_are_saved_as_literal_text_not_formulas(self) -> None:
        """Writing a leading equals sign as a formula would permit workbook injection."""
        records = pd.DataFrame(
            [
                {
                    "source": "EchoTik",
                    "name": '=HYPERLINK("https://attacker.invalid","click")',
                    "name_cn": "=1+1",
                    "gmv_7d": 1,
                    "detail_url": "https://echotik.example/product/1",
                },
                {
                    "source": "Amazon",
                    "name": "Complete Amazon product title",
                },
            ]
        )

        with patch(
            "scripts.ecommerce_report.workbook.translate_amazon_title_to_chinese",
            return_value="完整亚马逊商品标题",
        ):
            result = write_report(records, self.output_path, self.template_path)

        workbook = load_workbook(result, data_only=False)
        try:
            worksheet = workbook.active
            self.assertEqual(worksheet["D2"].value, '=HYPERLINK("https://attacker.invalid","click")')
            self.assertEqual(worksheet["D2"].data_type, "s")
            self.assertEqual(worksheet["E2"].value, "=1+1")
            self.assertEqual(worksheet["E2"].data_type, "s")
        finally:
            workbook.close()

    def test_a_new_formula_found_after_save_never_replaces_an_existing_report(self) -> None:
        """Removing the post-save formula gate would publish a corrupted workbook."""
        self.output_path.parent.mkdir(parents=True)
        original_report = b"existing valid report"
        self.output_path.write_bytes(original_report)
        original_save = Workbook.save

        def save_then_inject_formula(workbook, filename) -> None:
            original_save(workbook, filename)
            injected = load_workbook(filename)
            injected.active["D2"] = "=WEBSERVICE(\"https://attacker.invalid\")"
            original_save(injected, filename)
            injected.close()

        with patch.object(Workbook, "save", save_then_inject_formula):
            with patch(
                "scripts.ecommerce_report.workbook.translate_amazon_title_to_chinese",
                return_value="包含每个重要细节的完整英文商品标题",
            ):
                with self.assertRaisesRegex(ValueError, "模板之外的公式"):
                    write_report(
                        normalized_records(), self.output_path, self.template_path
                    )

        self.assertEqual(self.output_path.read_bytes(), original_report)

    def test_inspection_conservatively_reports_formulas_and_literal_errors_on_all_sheets(self) -> None:
        workbook = load_workbook(self.template_path)
        audit_sheet = workbook.create_sheet("审计")
        audit_sheet["A1"] = "=1/0"
        audit_sheet["A2"] = "#DIV/0!"
        workbook.save(self.template_path)
        workbook.close()

        inspection = inspect_report(self.template_path)

        self.assertEqual(
            inspection.formula_errors,
            ("审计!A1:=1/0", "审计!A2:#DIV/0!"),
        )


if __name__ == "__main__":
    unittest.main()
