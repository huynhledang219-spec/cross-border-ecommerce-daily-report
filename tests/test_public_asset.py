from __future__ import annotations

import hashlib
import posixpath
import re
import subprocess
import unittest
import xml.etree.ElementTree as ET
from collections import Counter
from pathlib import Path, PurePosixPath
from tempfile import TemporaryDirectory
from zipfile import ZipFile

import pandas as pd
import yaml
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from scripts.ecommerce_report.workbook import REPORT_HEADERS, write_report


REPOSITORY_ROOT = Path(__file__).resolve().parents[1]
ASSET_PATH = REPOSITORY_ROOT / "assets" / "report-template.xlsx"
APPROVED_HEADERS = (
    "排名",
    "近7天重点选品",
    "来源",
    "品名关键词",
    "中文名称",
    "价格(USD)",
    "商品评分",
    "评论数",
    "GMV",
    "7天GMV",
    "7天销量",
    "关联视频",
    "关联达人",
    "商品详情链接",
    "诊断",
)
EXPECTED_WIDTHS = (
    # The WPS source stores F:G and H:J as grouped OOXML column spans.
    # These are their effective rendered widths, not openpyxl's default for
    # a non-leading member of a grouped span.
    6.0,
    10.0,
    28.0,
    24.0,
    32.8571428571429,
    10.0,
    10.0,
    14.0,
    14.0,
    14.0,
    10.0,
    13.0,
    13.0,
    13.0,
    52.9333333333333,
)
EXPECTED_CHART_EXTENT = (3_513_455, 1_031_240)
PUBLIC_TEXT_SUFFIXES = frozenset({".md", ".yaml", ".yml", ".py", ".ps1", ".txt"})
PRIVATE_ABSOLUTE_PATH_PATTERNS = (
    re.compile(r"(?i)(?<![A-Za-z0-9])[A-Z]:[\\/]Users[\\/][^\\/\s\"'`]+"),
    re.compile(
        r"(?i)(?<![A-Za-z0-9])[A-Z]:[\\/][^\r\n\"'`]*(?:workspace|worktrees?|工作目录)"
    ),
    re.compile(r"(?i)(?<![A-Za-z0-9])/(?:home|Users)/[^/\s\"'`]+/"),
)
EXPECTED_PARTS = frozenset(
    {
        "[Content_Types].xml",
        "_rels/.rels",
        "docProps/app.xml",
        "docProps/core.xml",
        "xl/_rels/workbook.xml.rels",
        "xl/charts/chart1.xml",
        "xl/drawings/_rels/drawing1.xml.rels",
        "xl/drawings/drawing1.xml",
        "xl/styles.xml",
        "xl/theme/theme1.xml",
        "xl/workbook.xml",
        "xl/worksheets/_rels/sheet1.xml.rels",
        "xl/worksheets/sheet1.xml",
    }
)
EXPECTED_DEFAULT_CONTENT_TYPES = frozenset(
    {
        ("rels", "application/vnd.openxmlformats-package.relationships+xml"),
        ("xml", "application/xml"),
    }
)
EXPECTED_OVERRIDE_CONTENT_TYPES = {
    "/docProps/app.xml": (
        "application/vnd.openxmlformats-officedocument.extended-properties+xml"
    ),
    "/docProps/core.xml": "application/vnd.openxmlformats-package.core-properties+xml",
    "/xl/styles.xml": "application/vnd.openxmlformats-officedocument.spreadsheetml.styles+xml",
    "/xl/theme/theme1.xml": "application/vnd.openxmlformats-officedocument.theme+xml",
    "/xl/worksheets/sheet1.xml": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"
    ),
    "/xl/drawings/drawing1.xml": "application/vnd.openxmlformats-officedocument.drawing+xml",
    "/xl/charts/chart1.xml": "application/vnd.openxmlformats-officedocument.drawingml.chart+xml",
    "/xl/workbook.xml": (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"
    ),
}
EXPECTED_UNCHANGED_PART_HASHES = {
    "[Content_Types].xml": "63cf0c3306f3bf41b32d5ce2f43b3e5eed98b2eee6acd6a57e1a6ee524527893",
    "_rels/.rels": "c545941ba36c15fcdce4ae4568c663f3ced1a2226ad5082d3fd66b178bfac11a",
    "docProps/app.xml": "209fca6b00afe72a5029754b94be5953d8f16d96f67130325566b9366ad4ccc5",
    "docProps/core.xml": "185e39669d66701b6c580a68388bf4fee9eb5831bae3d95221739d42575b1ad0",
    "xl/_rels/workbook.xml.rels": "26ad8fcc38d41229833e624496df364492772697ad2e5d6696e1738f05ba225f",
    "xl/charts/chart1.xml": "cbec71313d8f59b369c05aba0146fafc9a63a9ae152be6cc15dc4653b1cb37c6",
    "xl/drawings/_rels/drawing1.xml.rels": "74667459a717caf5ced763d00fd999c3b180fd96c5804b901a08b8ecd5795128",
    "xl/drawings/drawing1.xml": "4fd8708453d71c978a6c8482e18c1271b3dbeb5ac37d22a035d654a0b1de6ccd",
    "xl/styles.xml": "f88d0db65c1cb0cbdaff569b8b1cbe6e2083222727edd4d4296bbc14f71be5cc",
    "xl/theme/theme1.xml": "a776ae573b94ab1d2aadb33fe62fd21dcae120ccd56d00139d58ff139040c6b3",
    "xl/workbook.xml": "698333647b55bdf82e43fdb4530e49a358bad7542c081baa8db924cde02cfc34",
    "xl/worksheets/_rels/sheet1.xml.rels": "0043881ca7f803da685c9fc201bc7c04f045356813c7b94a5d00253d3b5df9ee",
}


def _assert_asset_exists(test: unittest.TestCase) -> None:
    test.assertTrue(ASSET_PATH.is_file(), "public workbook asset is absent")


def _local_name(tag: str) -> str:
    return tag.rsplit("}", 1)[-1]


def _manifest_is_allowlisted(member_names: list[str]) -> bool:
    unique_names = frozenset(member_names)
    return all(
        (
            len(member_names) == 13,
            len(member_names) == len(unique_names),
            unique_names == EXPECTED_PARTS,
        )
    )


def _relationship_target(rels_part: str, target: str) -> str:
    if rels_part == "_rels/.rels":
        base = PurePosixPath()
    else:
        rels_path = PurePosixPath(rels_part)
        source_name = rels_path.name.removesuffix(".rels")
        source_part = rels_path.parent.parent / source_name
        base = source_part.parent
    if target.startswith("/"):
        candidate = target.lstrip("/")
    else:
        candidate = str(base / target)
    return posixpath.normpath(candidate)


def _header_style_is_approved(cell) -> bool:
    font_color = cell.font.color
    return all(
        (
            cell.style_id == 2,
            cell.font.name == "微软雅黑",
            cell.font.sz == 12.0,
            cell.font.bold is True,
            cell.font.italic is False,
            font_color is not None,
            font_color.type == "rgb" if font_color else False,
            font_color.rgb == "FF000000" if font_color else False,
            cell.fill.fill_type == "solid",
            cell.fill.fgColor.type == "rgb",
            cell.fill.fgColor.rgb == "FFFFD700",
            cell.border.left.style == "thin",
            cell.border.right.style == "thin",
            cell.border.top.style == "thin",
            cell.border.bottom.style == "thin",
            cell.alignment.horizontal == "center",
            cell.alignment.vertical == "center",
            cell.alignment.wrap_text is True,
            cell.number_format == "General",
            cell.protection.locked is True,
            cell.protection.hidden is False,
        )
    )


class PublicWorkbookAssetTests(unittest.TestCase):
    def test_hidden_header_migration_preserves_every_other_zip_part_exactly(
        self,
    ) -> None:
        _assert_asset_exists(self)
        with ZipFile(ASSET_PATH) as archive:
            actual = {
                name: hashlib.sha256(archive.read(name)).hexdigest()
                for name in EXPECTED_UNCHANGED_PART_HASHES
            }
        self.assertEqual(actual, EXPECTED_UNCHANGED_PART_HASHES)

    def test_preserves_the_actual_2026_8_11_layout_without_business_values(self) -> None:
        """The public asset must be a sanitized layout clone, not a header-only shell."""

        _assert_asset_exists(self)
        workbook = load_workbook(ASSET_PATH, read_only=False, data_only=False, keep_links=False)
        try:
            worksheet = workbook.active
            self.assertEqual(worksheet.title, "Sheet1")
            self.assertGreaterEqual(worksheet.max_row, 4)
            self.assertGreaterEqual(worksheet.max_column, 22)
            self.assertEqual(worksheet.freeze_panes, "A2")
            self.assertEqual(worksheet.auto_filter.ref, "A1:O4")
            self.assertEqual(
                tuple(
                    worksheet.column_dimensions[get_column_letter(column)].width
                    for column in range(1, 16)
                ),
                EXPECTED_WIDTHS,
            )
            self.assertEqual(worksheet.row_dimensions[1].height, 32.0)
            self.assertEqual(worksheet.row_dimensions[2].height, 78.0)
            self.assertEqual(worksheet.row_dimensions[3].height, 78.0)
            self.assertEqual(worksheet.row_dimensions[4].height, 78.0)
            self.assertTrue(worksheet.column_dimensions["N"].hidden)
            self.assertTrue(all(worksheet.column_dimensions[column].hidden for column in "PQRSTUV"))
            self.assertEqual(worksheet["A2"].fill.fgColor.rgb, "FFE8F5E9")
            self.assertEqual(worksheet["C2"].alignment.horizontal, "left")
            self.assertEqual(worksheet["D3"].alignment.horizontal, "left")
            self.assertEqual(worksheet["E3"].alignment.horizontal, "left")
            self.assertNotEqual(worksheet["A3"].fill.fgColor.rgb, worksheet["A4"].fill.fgColor.rgb)
            self.assertEqual(len(worksheet._charts), 1)
            chart = worksheet._charts[0]
            self.assertEqual(len(chart.series), 0)
            self.assertEqual((chart.anchor._from.col, chart.anchor._from.row), (14, 2))
            self.assertEqual((chart.anchor.ext.cx, chart.anchor.ext.cy), EXPECTED_CHART_EXTENT)

            visible_values = tuple(
                cell.value
                for row in worksheet.iter_rows()
                for cell in row
                if cell.value is not None
            )
            self.assertEqual(visible_values, APPROVED_HEADERS)
        finally:
            workbook.close()

    def test_manifest_check_rejects_a_duplicate_allowlisted_part(self) -> None:
        member_names = list(EXPECTED_PARTS)
        member_names.append("xl/workbook.xml")
        self.assertFalse(
            _manifest_is_allowlisted(member_names),
            "ZIP manifest check accepted a duplicate allowlisted part",
        )

    def test_contains_exactly_the_approved_visible_text(self) -> None:
        _assert_asset_exists(self)
        self.assertTrue(
            tuple(REPORT_HEADERS) == APPROVED_HEADERS,
            "report writer headers differ from the public asset contract",
        )
        workbook = load_workbook(ASSET_PATH, read_only=False, data_only=False, keep_links=False)
        try:
            self.assertEqual(len(workbook.worksheets), 1)
            worksheet = workbook.active
            self.assertEqual(worksheet.sheet_state, "visible")
            self.assertTrue(worksheet.title == "Sheet1", "worksheet title is not allowlisted")
            self.assertEqual(worksheet.max_row, 4)
            self.assertEqual(worksheet.max_column, 22)
            visible_values = tuple(
                cell.value
                for row in worksheet.iter_rows()
                for cell in row
                if cell.value is not None
            )
            self.assertTrue(
                visible_values == APPROVED_HEADERS,
                "worksheet visible text differs from the approved 15-header allowlist",
            )
        finally:
            workbook.close()

    def test_retains_static_layout_style_and_empty_chart_contract(self) -> None:
        _assert_asset_exists(self)
        workbook = load_workbook(ASSET_PATH, read_only=False, data_only=False, keep_links=False)
        try:
            worksheet = workbook.active
            widths = tuple(
                worksheet.column_dimensions[get_column_letter(column)].width
                for column in range(1, 16)
            )
            self.assertTrue(widths == EXPECTED_WIDTHS, "column widths differ from the public contract")
            self.assertEqual(worksheet.row_dimensions[1].height, 32.0)
            self.assertEqual(worksheet.row_dimensions[2].height, 78.0)
            self.assertEqual(worksheet.row_dimensions[3].height, 78.0)
            self.assertEqual(worksheet.row_dimensions[4].height, 78.0)
            self.assertTrue(
                all(_header_style_is_approved(worksheet.cell(1, column)) for column in range(1, 16)),
                "header style differs from the public contract",
            )
            self.assertEqual(len(worksheet._charts), 1)
            chart = worksheet._charts[0]
            self.assertEqual(len(chart.series), 0)
            self.assertEqual((chart.anchor.ext.cx, chart.anchor.ext.cy), EXPECTED_CHART_EXTENT)
        finally:
            workbook.close()

    def test_has_no_workbook_or_cell_active_content(self) -> None:
        _assert_asset_exists(self)
        workbook = load_workbook(ASSET_PATH, read_only=False, data_only=False, keep_links=False)
        try:
            self.assertTrue(
                not list(workbook.defined_names.values()), "workbook contains defined names"
            )
            self.assertTrue(
                not bool(getattr(workbook, "_external_links", [])),
                "workbook has external links",
            )
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows():
                    for cell in row:
                        self.assertTrue(cell.data_type != "f", "worksheet contains a formula")
                        self.assertTrue(cell.hyperlink is None, "worksheet contains a hyperlink")
                        self.assertTrue(cell.comment is None, "worksheet contains a comment")
        finally:
            workbook.close()

    def test_zip_manifest_content_types_relationships_and_text_are_allowlisted(self) -> None:
        _assert_asset_exists(self)
        with ZipFile(ASSET_PATH) as archive:
            member_names = archive.namelist()
            self.assertTrue(
                _manifest_is_allowlisted(member_names),
                "ZIP manifest differs from the exact allowlist",
            )
            names = frozenset(member_names)
            self.assertNotIn("docProps/custom.xml", names, "ZIP contains private custom properties")

            core_properties = ET.fromstring(archive.read("docProps/core.xml"))
            core_text = {
                _local_name(element.tag): (element.text or "").strip()
                for element in core_properties
            }
            self.assertEqual(core_text.get("creator"), "")
            self.assertEqual(core_text.get("lastModifiedBy"), "")
            self.assertEqual(core_text.get("created"), "2000-01-01T00:00:00Z")
            app_properties = ET.fromstring(archive.read("docProps/app.xml"))
            app_text = {
                _local_name(element.tag): (element.text or "").strip()
                for element in app_properties
            }
            self.assertEqual(app_text.get("Application"), "Microsoft Excel Compatible / Openpyxl 3.1.5")

            content_types = ET.fromstring(archive.read("[Content_Types].xml"))
            defaults = frozenset(
                (element.attrib.get("Extension"), element.attrib.get("ContentType"))
                for element in content_types
                if _local_name(element.tag) == "Default"
            )
            overrides = {
                element.attrib.get("PartName"): element.attrib.get("ContentType")
                for element in content_types
                if _local_name(element.tag) == "Override"
            }
            self.assertTrue(
                defaults == EXPECTED_DEFAULT_CONTENT_TYPES,
                "default Content Types differ from the exact allowlist",
            )
            self.assertTrue(
                overrides == EXPECTED_OVERRIDE_CONTENT_TYPES,
                "override Content Types differ from the exact allowlist",
            )

            visible_xml_text: list[str] = []
            for name in sorted(names):
                if not name.endswith((".xml", ".rels")):
                    continue
                root = ET.fromstring(archive.read(name))
                for element in root.iter():
                    local_name = _local_name(element.tag)
                    self.assertTrue(local_name != "f", "package XML contains a formula reference")
                    if local_name == "t" and element.text:
                        visible_xml_text.append(element.text)
                if not name.endswith(".rels"):
                    continue
                for relationship in root:
                    self.assertTrue(
                        relationship.attrib.get("TargetMode", "").lower() != "external",
                        "package contains an external relationship",
                    )
                    target = relationship.attrib.get("Target", "")
                    self.assertTrue(bool(target), "package relationship has no target")
                    resolved_target = _relationship_target(name, target)
                    self.assertTrue(
                        resolved_target in names,
                        "package relationship target is outside the exact manifest allowlist",
                    )

            self.assertTrue(
                Counter(visible_xml_text) == Counter(APPROVED_HEADERS),
                "package visible text differs from the approved 15-header allowlist",
            )

            for name in names - {"[Content_Types].xml"}:
                if name.endswith(".rels"):
                    content_type = dict(defaults).get("rels")
                else:
                    content_type = overrides.get(f"/{name}") or dict(defaults).get(
                        PurePosixPath(name).suffix.lstrip(".")
                    )
                self.assertTrue(bool(content_type), "manifest part has no allowlisted Content Type")

    def test_write_report_can_consume_the_public_asset(self) -> None:
        _assert_asset_exists(self)
        with TemporaryDirectory() as temporary_directory:
            output_path = Path(temporary_directory) / "report.xlsx"
            result = write_report(pd.DataFrame(), output_path, ASSET_PATH)
            self.assertEqual(result, output_path)
            workbook = load_workbook(output_path, read_only=False, data_only=False, keep_links=False)
            workbook.close()
            self.assertTrue(output_path.is_file(), "write_report did not create an output workbook")

    def test_public_asset_generates_a_nonempty_report_with_actual_row_banding(self) -> None:
        _assert_asset_exists(self)
        records = pd.DataFrame(
            [
                {
                    "rank": "SKU",
                    "source": "你的库存",
                    "name": "Inventory item",
                    "name_cn": "库存商品",
                    "price": 12.0,
                    "diagnostic": "库存诊断",
                },
                *[
                    {
                        "rank": index,
                        "source": "EchoTik",
                        "name": f"Product {index}",
                        "name_cn": f"商品 {index}",
                        "price": 10.0 + index,
                        "gmv": 1000 - index,
                        "gmv_7d": 100 - index,
                        "gmv_trend_7d": [1, 2, 3, 4, 5, 6, 7 + index],
                    }
                    for index in range(1, 4)
                ],
            ]
        )
        with TemporaryDirectory() as temporary_directory:
            output_path = Path(temporary_directory) / "report.xlsx"
            write_report(records, output_path, ASSET_PATH)
            workbook = load_workbook(output_path, read_only=False, data_only=False, keep_links=False)
            try:
                worksheet = workbook.active
                self.assertEqual(worksheet["A2"].fill.fgColor.rgb, "FFE8F5E9")
                self.assertNotEqual(worksheet["A3"].fill.fgColor.rgb, worksheet["A4"].fill.fgColor.rgb)
                self.assertEqual(worksheet["A5"].fill.fgColor.rgb, worksheet["A3"].fill.fgColor.rgb)
                self.assertEqual(len(worksheet._charts), 3)
                self.assertTrue(
                    all(
                        (chart.anchor.ext.cx, chart.anchor.ext.cy) == EXPECTED_CHART_EXTENT
                        for chart in worksheet._charts
                    )
                )
            finally:
                workbook.close()


class PublicSkillGuidanceTests(unittest.TestCase):
    def test_tracked_public_text_has_no_literal_private_absolute_paths(self) -> None:
        portable_examples = (
            '$codexHome = $env:CODEX_HOME',
            '$codexHome = Join-Path $HOME ".codex"',
            'python (Join-Path $codexHome "skills/.system/skill-creator/scripts/quick_validate.py") "."',
            'adapter: C:/temp/adapter.py',
        )
        self.assertFalse(
            any(
                pattern.search(example)
                for example in portable_examples
                for pattern in PRIVATE_ABSOLUTE_PATH_PATTERNS
            ),
            "portable environment-based path syntax was rejected",
        )

        tracked = subprocess.run(
            ["git", "ls-files", "-z"],
            cwd=REPOSITORY_ROOT,
            check=True,
            capture_output=True,
        ).stdout.decode("utf-8").split("\0")
        findings: list[str] = []
        for relative_name in tracked:
            if not relative_name:
                continue
            relative_path = Path(relative_name)
            # Test fixtures intentionally exercise sanitized, fictional profile paths.
            if relative_path.parts[0] == "tests":
                continue
            if relative_path.suffix.lower() not in PUBLIC_TEXT_SUFFIXES:
                continue
            content = (REPOSITORY_ROOT / relative_path).read_text(encoding="utf-8")
            if any(pattern.search(content) for pattern in PRIVATE_ABSOLUTE_PATH_PATTERNS):
                findings.append(relative_path.as_posix())

        self.assertEqual(
            findings,
            [],
            "tracked public text contains literal private absolute paths in: "
            + ", ".join(findings),
        )

    def test_public_guidance_documents_default_and_replaceable_platform(self) -> None:
        skill = (REPOSITORY_ROOT / "SKILL.md").read_text(encoding="utf-8")
        config_reference = (
            REPOSITORY_ROOT / "references" / "configuration.md"
        ).read_text(encoding="utf-8")
        report_schema = (
            REPOSITORY_ROOT / "references" / "report-schema.md"
        ).read_text(encoding="utf-8")
        config_example = (
            REPOSITORY_ROOT / "scripts" / "config.example.yaml"
        ).read_text(encoding="utf-8")

        self.assertIn("EchoTik remains the default", skill)
        self.assertIn("Amazon remains a required supplementary source", skill)
        self.assertIn("equivalent-capability gate", skill)
        self.assertIn("Naming a website does not make it compatible", skill)
        self.assertIn("registered adapter", config_reference)
        self.assertIn("closed adapter key", config_reference)
        self.assertIn("never imports an executable path or remote code", config_reference)
        self.assertIn("Human verification", config_reference)
        self.assertIn("normalized record contract", report_schema)
        self.assertIn("unchanged across registered adapters", report_schema)
        self.assertIn("primary_platform:", config_example)
        self.assertIn("adapter: echotik", config_example)
        self.assertNotIn("README.md", {path.name for path in REPOSITORY_ROOT.iterdir()})

    def test_skill_frontmatter_uses_the_approved_platform_aware_trigger(self) -> None:
        skill = (REPOSITORY_ROOT / "SKILL.md").read_text(encoding="utf-8")
        expected = (
            "---\n"
            "name: cross-border-ecommerce-daily-report\n"
            "description: Use when configuring, running, scheduling, troubleshooting, or "
            "validating a Windows daily product-intelligence report that uses EchoTik by "
            "default or a verified registered platform adapter, with Amazon as a "
            "supplementary source.\n"
            "---"
        )
        self.assertTrue(skill.startswith(expected))

    def test_openai_metadata_is_fully_english_and_exact(self) -> None:
        metadata = yaml.safe_load(
            (REPOSITORY_ROOT / "agents" / "openai.yaml").read_text(encoding="utf-8")
        )
        serialized = yaml.safe_dump(metadata, allow_unicode=True)
        self.assertIsNone(re.search(r"[\u3400-\u9fff]", serialized))
        self.assertEqual(
            metadata,
            {
                "interface": {
                    "display_name": "Cross-Border E-Commerce Daily Report",
                    "short_description": (
                        "Generate validated multi-source product-intelligence reports"
                    ),
                    "default_prompt": (
                        "Use $cross-border-ecommerce-daily-report to configure, generate, "
                        "and verify today's cross-border e-commerce product-intelligence "
                        "report."
                    ),
                }
            },
        )


if __name__ == "__main__":
    unittest.main()
