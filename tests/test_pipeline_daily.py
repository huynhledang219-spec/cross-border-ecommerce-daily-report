from __future__ import annotations

import io
import os
import subprocess
import unittest
from contextlib import redirect_stderr
from datetime import date, datetime
from pathlib import Path
from tempfile import TemporaryDirectory
from unittest.mock import Mock, patch

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference

from scripts.ecommerce_report.config import RuntimeConfig
from scripts.ecommerce_report import daily as daily_module
from scripts.ecommerce_report import pipeline as pipeline_module
from scripts.ecommerce_report import workbook as workbook_module
from scripts.ecommerce_report.daily import failure_path_for, run_daily_job
from scripts.ecommerce_report.pipeline import PipelineError, run_pipeline
from scripts.ecommerce_report.platforms import (
    PlatformAdapterRegistry,
    PlatformCapabilities,
    PrimaryPlatformConfig,
)
from scripts.ecommerce_report.workbook import REPORT_HEADERS
from scripts.run_daily import main as run_daily_main
from scripts.run_report import main as run_report_main


class _PlaywrightSession:
    def __init__(
        self,
        enter_error: BaseException | None = None,
        exit_error: BaseException | None = None,
    ) -> None:
        self.enter_error = enter_error
        self.exit_error = exit_error

    def __enter__(self):
        if self.enter_error is not None:
            raise self.enter_error
        return object()

    def __exit__(self, exc_type, exc_value, traceback) -> None:
        if self.exit_error is not None:
            raise self.exit_error
        return None


class _RecordingAdapter:
    key = "marketpulse"
    display_name = "MarketPulse"
    capabilities = PlatformCapabilities(True, True, True, True)

    def __init__(self, records: pd.DataFrame) -> None:
        self.records = records
        self.collect_calls: list[dict[str, object]] = []
        self.collection_error: BaseException | None = None

    def validate_config(self, config: PrimaryPlatformConfig) -> None:
        if config.adapter != self.key or not config.categories:
            raise ValueError("MarketPulse requires configured categories")

    def collect(
        self,
        context,
        config: PrimaryPlatformConfig,
        *,
        detail_limit: int,
        trend_days: int,
        pages_per_category: int,
    ) -> pd.DataFrame:
        self.collect_calls.append(
            {
                "context": context,
                "config": config,
                "detail_limit": detail_limit,
                "trend_days": trend_days,
                "pages_per_category": pages_per_category,
            }
        )
        if self.collection_error is not None:
            raise self.collection_error
        return self.records


def _complete_platform_dataframe(
    source: str = "MarketPulse", count: int = 1
) -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "source": source,
                "name": f"Product {index}",
                "name_cn": f"Product CN {index}",
                "category": "Home / Kitchen",
                "price": 10.0,
                "rating": 4.5,
                "reviews": 100,
                "gmv": 1000.0,
                "gmv_7d": 700.0 - index,
                "sold_7d": 70,
                "videos": 7,
                "creators": 3,
                "detail_url": f"https://example.test/products/{index}",
                "gmv_trend_7d": [100.0] * 7,
            }
            for index in range(count)
        ]
    )


def _complete_amazon_dataframe() -> pd.DataFrame:
    return pd.DataFrame([{"source": "Amazon", "name": "Amazon product"}])


def _create_template(path: Path) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(REPORT_HEADERS)
    worksheet.append([None] * len(REPORT_HEADERS))
    for column, value in enumerate(range(1, 8), start=16):
        worksheet.cell(2, column, value)
    chart = LineChart()
    chart.width = 3_513_455 / 360_000
    chart.height = 1_031_240 / 360_000
    chart.add_data(
        Reference(worksheet, min_col=16, max_col=22, min_row=2, max_row=2),
        from_rows=True,
    )
    worksheet.add_chart(chart, "O2")
    workbook.save(path)
    workbook.close()


class PipelineTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.template_path = self.root / "template.xlsx"
        self.output_path = self.root / "reports" / "report.xlsx"
        _create_template(self.template_path)
        self.config = RuntimeConfig.from_mapping(
            self.root,
            {
                "output_dir": "./reports",
                "profile_dir": "./profile",
                "template_path": "./template.xlsx",
                "echotik_categories": [
                    {"path": ["Home", "Kitchen"], "id": "123456"},
                ],
                "amazon_categories": [
                    {"name": "Kitchen", "url": "https://www.amazon.com/kitchen"},
                ],
            },
        )

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def _marketpulse_config(self) -> RuntimeConfig:
        return RuntimeConfig(
            output_dir=self.config.output_dir,
            profile_dir=self.config.profile_dir,
            template_path=self.config.template_path,
            detail_limit=20,
            trend_days=7,
            pages_per_category=4,
            primary_platform=PrimaryPlatformConfig(
                adapter="marketpulse",
                categories=({"path": ["Home", "Kitchen"], "id": "42"},),
            ),
            amazon_categories=self.config.amazon_categories,
        )

    def test_pipeline_uses_the_native_platform_neutral_workbook_writer(self) -> None:
        self.assertIs(pipeline_module.write_report, workbook_module.write_report)

    def test_pipeline_runs_registered_non_echotik_adapter(self) -> None:
        adapter = _RecordingAdapter(_complete_platform_dataframe(count=21))
        registry = PlatformAdapterRegistry((adapter,))
        config = self._marketpulse_config()
        browser_context = Mock()

        with patch.object(
            pipeline_module, "_playwright_session", return_value=_PlaywrightSession()
        ), patch.object(
            pipeline_module,
            "open_platform_context",
            return_value=browser_context,
            create=True,
        ), patch.object(
            pipeline_module,
            "scrape_amazon",
            return_value=_complete_amazon_dataframe(),
        ), patch.object(
            pipeline_module, "write_report", return_value=self.output_path
        ) as write_report_mock:
            result = run_pipeline(config, self.output_path, registry=registry)

        self.assertEqual(result, self.output_path)
        self.assertEqual(len(adapter.collect_calls), 1)
        self.assertEqual(adapter.collect_calls[0]["detail_limit"], 20)
        self.assertEqual(adapter.collect_calls[0]["trend_days"], 7)
        self.assertEqual(adapter.collect_calls[0]["pages_per_category"], 4)
        self.assertEqual(
            write_report_mock.call_args.kwargs["primary_source"], "MarketPulse"
        )
        browser_context.close.assert_called_once_with()

    def test_non_echotik_collection_failure_uses_dynamic_stage(self) -> None:
        adapter = _RecordingAdapter(_complete_platform_dataframe())
        adapter.collection_error = RuntimeError("service unavailable")
        registry = PlatformAdapterRegistry((adapter,))

        with patch.object(
            pipeline_module, "_playwright_session", return_value=_PlaywrightSession()
        ), patch.object(
            pipeline_module, "open_platform_context", return_value=Mock(), create=True
        ), patch.object(pipeline_module, "scrape_amazon") as scrape_amazon:
            with self.assertRaises(PipelineError) as raised:
                run_pipeline(
                    self._marketpulse_config(), self.output_path, registry=registry
                )

        self.assertEqual(raised.exception.stage, "MarketPulse采集")
        self.assertEqual(str(raised.exception.error), "service unavailable")
        scrape_amazon.assert_not_called()

    def test_empty_non_echotik_records_stop_before_export(self) -> None:
        adapter = _RecordingAdapter(pd.DataFrame())
        registry = PlatformAdapterRegistry((adapter,))

        with patch.object(
            pipeline_module, "_playwright_session", return_value=_PlaywrightSession()
        ), patch.object(
            pipeline_module, "open_platform_context", return_value=Mock(), create=True
        ), patch.object(
            pipeline_module,
            "scrape_amazon",
            return_value=_complete_amazon_dataframe(),
        ), patch.object(pipeline_module, "write_report") as write_report_mock:
            with self.assertRaises(PipelineError) as raised:
                run_pipeline(
                    self._marketpulse_config(), self.output_path, registry=registry
                )

        self.assertEqual(raised.exception.stage, "MarketPulse采集")
        write_report_mock.assert_not_called()

    def test_invalid_normalized_records_stop_before_export(self) -> None:
        adapter = _RecordingAdapter(
            pd.DataFrame([{"source": "MarketPulse", "name": "Incomplete"}])
        )
        registry = PlatformAdapterRegistry((adapter,))

        with patch.object(
            pipeline_module, "_playwright_session", return_value=_PlaywrightSession()
        ), patch.object(
            pipeline_module, "open_platform_context", return_value=Mock(), create=True
        ), patch.object(pipeline_module, "scrape_amazon") as scrape_amazon, patch.object(
            pipeline_module, "write_report"
        ) as write_report_mock:
            with self.assertRaises(PipelineError) as raised:
                run_pipeline(
                    self._marketpulse_config(), self.output_path, registry=registry
                )

        self.assertEqual(raised.exception.stage, "MarketPulse采集")
        self.assertIn("normalized product fields", str(raised.exception.error))
        scrape_amazon.assert_called_once()
        write_report_mock.assert_not_called()

    def test_malformed_trend_stops_before_export_instead_of_becoming_data_empty(self) -> None:
        records = _complete_platform_dataframe()
        records.at[0, "gmv_trend_7d"] = [1, 2, 3, 4, 5, 6, float("nan")]
        adapter = _RecordingAdapter(records)
        registry = PlatformAdapterRegistry((adapter,))

        with patch.object(
            pipeline_module, "_playwright_session", return_value=_PlaywrightSession()
        ), patch.object(
            pipeline_module, "open_platform_context", return_value=Mock(), create=True
        ), patch.object(
            pipeline_module, "scrape_amazon", return_value=_complete_amazon_dataframe()
        ), patch.object(pipeline_module, "write_report") as write_report_mock:
            with self.assertRaises(PipelineError) as raised:
                run_pipeline(
                    self._marketpulse_config(), self.output_path, registry=registry
                )

        self.assertEqual(raised.exception.stage, "MarketPulse采集")
        self.assertIn("exactly seven finite nonnegative", str(raised.exception.error))
        write_report_mock.assert_not_called()

    @patch("scripts.ecommerce_report.pipeline.scrape_amazon")
    @patch("scripts.ecommerce_report.echotik.ECHOTIK_ADAPTER.collect")
    @patch("scripts.ecommerce_report.pipeline.open_platform_context")
    @patch("scripts.ecommerce_report.pipeline._playwright_session")
    @patch(
        "scripts.ecommerce_report.workbook.translate_amazon_title_to_chinese",
        return_value="亚马逊完整商品标题",
    )
    def test_an_empty_top_trend_does_not_block_a_later_top_twenty_chart(
        self,
        translate_amazon: Mock,
        playwright_session: Mock,
        open_context: Mock,
        scrape_echotik: Mock,
        scrape_amazon: Mock,
    ) -> None:
        """Aborting export after one empty trend would lose the later valid chart."""
        playwright_session.return_value = _PlaywrightSession()
        browser_context = Mock()
        open_context.return_value = browser_context
        echotik_records = _complete_platform_dataframe("EchoTik", count=2)
        echotik_records.loc[0, "name"] = "empty trend"
        echotik_records.loc[0, "gmv"] = 1000
        echotik_records.loc[0, "gmv_7d"] = 100
        echotik_records.at[0, "gmv_trend_7d"] = None
        echotik_records.loc[1, "name"] = "later trend"
        echotik_records.loc[1, "gmv"] = 900
        echotik_records.loc[1, "gmv_7d"] = 90
        echotik_records.at[1, "gmv_trend_7d"] = [1, 2, 3, 4, 5, 6, 7]
        scrape_echotik.return_value = echotik_records
        scrape_amazon.return_value = pd.DataFrame(
            [{"source": "Amazon", "name": "Complete Amazon product title"}]
        )

        result = run_pipeline(self.config, self.output_path)

        workbook = load_workbook(result, data_only=False)
        try:
            worksheet = workbook.active
            self.assertEqual(worksheet["B2"].value, "Top 1")
            self.assertEqual(worksheet["O2"].value, "数据为空")
            self.assertEqual(worksheet["B3"].value, "Top 2")
            self.assertIsNone(worksheet["O3"].value)
            self.assertEqual(len(worksheet._charts), 1)
        finally:
            workbook.close()
        browser_context.close.assert_called_once_with()

    @patch("scripts.ecommerce_report.pipeline._playwright_session")
    def test_playwright_enter_failure_is_labeled_as_browser_start(
        self, playwright_session: Mock
    ) -> None:
        """Leaving __enter__ unwrapped would lose the actionable startup stage."""
        playwright_session.return_value = _PlaywrightSession(
            enter_error=RuntimeError("enter failed")
        )

        with self.assertRaises(PipelineError) as raised:
            run_pipeline(self.config, self.output_path)

        self.assertEqual(raised.exception.stage, "启动浏览器")
        self.assertEqual(str(raised.exception.error), "enter failed")

    @patch("scripts.ecommerce_report.pipeline.scrape_amazon")
    @patch("scripts.ecommerce_report.echotik.ECHOTIK_ADAPTER.collect")
    @patch("scripts.ecommerce_report.pipeline.open_platform_context")
    @patch("scripts.ecommerce_report.pipeline._playwright_session")
    def test_close_failures_do_not_mask_an_echotik_collection_failure(
        self,
        playwright_session: Mock,
        open_context: Mock,
        scrape_echotik: Mock,
        scrape_amazon: Mock,
    ) -> None:
        """Replacing the source error with cleanup noise would report the wrong stage."""
        playwright_session.return_value = _PlaywrightSession(
            exit_error=RuntimeError("session exit failed")
        )
        browser_context = Mock()
        browser_context.close.side_effect = RuntimeError("context close failed")
        open_context.return_value = browser_context
        scrape_echotik.side_effect = RuntimeError("EchoTik unavailable")

        with self.assertRaises(PipelineError) as raised:
            run_pipeline(self.config, self.output_path)

        self.assertEqual(raised.exception.stage, "EchoTik采集")
        self.assertEqual(str(raised.exception.error), "EchoTik unavailable")
        scrape_amazon.assert_not_called()

    @patch("scripts.ecommerce_report.pipeline.scrape_amazon")
    @patch("scripts.ecommerce_report.echotik.ECHOTIK_ADAPTER.collect")
    @patch("scripts.ecommerce_report.pipeline.open_platform_context")
    @patch("scripts.ecommerce_report.pipeline._playwright_session")
    def test_playwright_exit_failure_is_labeled_as_browser_close(
        self,
        playwright_session: Mock,
        open_context: Mock,
        scrape_echotik: Mock,
        scrape_amazon: Mock,
    ) -> None:
        """An exit failure after successful collection must identify cleanup as its stage."""
        playwright_session.return_value = _PlaywrightSession(
            exit_error=RuntimeError("session exit failed")
        )
        open_context.return_value = Mock()
        scrape_echotik.return_value = pd.DataFrame()
        scrape_amazon.return_value = pd.DataFrame()

        with self.assertRaises(PipelineError) as raised:
            run_pipeline(self.config, self.output_path)

        self.assertEqual(raised.exception.stage, "关闭浏览器")
        self.assertEqual(str(raised.exception.error), "session exit failed")

    @patch("scripts.ecommerce_report.pipeline.write_report")
    @patch("scripts.ecommerce_report.pipeline.scrape_amazon")
    @patch("scripts.ecommerce_report.echotik.ECHOTIK_ADAPTER.collect")
    @patch("scripts.ecommerce_report.pipeline.open_platform_context")
    @patch("scripts.ecommerce_report.pipeline._playwright_session")
    def test_amazon_and_export_failures_keep_their_exact_stages(
        self,
        playwright_session: Mock,
        open_context: Mock,
        scrape_echotik: Mock,
        scrape_amazon: Mock,
        write_report_mock: Mock,
    ) -> None:
        """A generic pipeline stage would not identify the failed boundary."""
        playwright_session.return_value = _PlaywrightSession()
        open_context.return_value = Mock()
        scrape_echotik.return_value = pd.DataFrame()
        scrape_amazon.side_effect = RuntimeError("Amazon unavailable")

        with self.assertRaises(PipelineError) as amazon_raised:
            run_pipeline(self.config, self.output_path)

        self.assertEqual(amazon_raised.exception.stage, "Amazon采集")

        scrape_amazon.side_effect = None
        scrape_echotik.return_value = _complete_platform_dataframe("EchoTik")
        scrape_amazon.return_value = pd.DataFrame(
            [{"source": "Amazon", "name": "Amazon item"}]
        )
        write_report_mock.side_effect = RuntimeError("workbook unavailable")

        with self.assertRaises(PipelineError) as export_raised:
            run_pipeline(self.config, self.output_path)

        self.assertEqual(export_raised.exception.stage, "导出报表")

    @patch("scripts.ecommerce_report.pipeline.write_report")
    @patch("scripts.ecommerce_report.pipeline.scrape_amazon")
    @patch("scripts.ecommerce_report.echotik.ECHOTIK_ADAPTER.collect")
    @patch("scripts.ecommerce_report.pipeline.open_platform_context")
    @patch("scripts.ecommerce_report.pipeline._playwright_session")
    def test_an_empty_required_source_never_exports_a_report(
        self,
        playwright_session: Mock,
        open_context: Mock,
        scrape_echotik: Mock,
        scrape_amazon: Mock,
        write_report_mock: Mock,
    ) -> None:
        """Concatenating whichever source is nonempty would hide a failed required source."""
        playwright_session.return_value = _PlaywrightSession()
        open_context.return_value = Mock()
        primary_row = _complete_platform_dataframe("EchoTik")
        amazon_row = _complete_amazon_dataframe()

        for empty_name, expected_stage in (
            ("echotik", "EchoTik采集"),
            ("amazon", "Amazon采集"),
        ):
            with self.subTest(empty_name=empty_name):
                scrape_echotik.return_value = (
                    pd.DataFrame() if empty_name == "echotik" else primary_row
                )
                scrape_amazon.return_value = (
                    pd.DataFrame() if empty_name == "amazon" else amazon_row
                )

                with self.assertRaises(PipelineError) as raised:
                    run_pipeline(self.config, self.output_path)

                self.assertEqual(raised.exception.stage, expected_stage)

        write_report_mock.assert_not_called()

    def test_pipeline_rejects_an_explicit_output_inside_the_skill(self) -> None:
        """A manual --output override must not bypass the runtime isolation policy."""
        skill_root = Path(__file__).resolve().parents[1]

        with patch("scripts.ecommerce_report.pipeline._playwright_session") as session:
            try:
                run_pipeline(self.config, skill_root / "generated-report.xlsx")
            except Exception as error:
                self.assertIs(type(error), ValueError)
                self.assertEqual(
                    str(error),
                    "output path must not be inside the Skill directory",
                )
            else:
                self.fail("Skill-local output was accepted")
        session.assert_not_called()


class DailyJobTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temp_dir = TemporaryDirectory()
        self.root = Path(self.temp_dir.name)
        self.config_path = self.root / "config.yaml"
        self.config_path.write_text("unused", encoding="utf-8")
        self.config = Mock(output_dir=self.root / "reports")
        self.day = date(2026, 8, 11)

    def tearDown(self) -> None:
        self.temp_dir.cleanup()

    def test_failure_path_uses_the_dedicated_daily_failure_directory(self) -> None:
        """Changing the directory or date format would hide failures from operators."""
        self.assertEqual(
            failure_path_for(self.day, self.root / "reports"),
            self.root / "reports" / "数据报表_失败原因" / "2026.8.11失败原因.txt",
        )

    @patch("scripts.ecommerce_report.daily.run_pipeline")
    @patch("scripts.ecommerce_report.daily.RuntimeConfig.load")
    def test_existing_daily_report_is_returned_without_collecting_again(
        self, load_config: Mock, run_pipeline_mock: Mock
    ) -> None:
        """Removing the existence guard would repeat a completed daily scrape."""
        load_config.return_value = self.config
        expected = self.config.output_dir / "2026.8.11数据报表.xlsx"
        expected.parent.mkdir(parents=True)
        expected.write_bytes(b"complete")

        result = run_daily_job(self.config_path, self.day)

        self.assertEqual(result, expected)
        run_pipeline_mock.assert_not_called()

    @patch("scripts.ecommerce_report.daily._now")
    @patch("scripts.ecommerce_report.daily.run_pipeline")
    @patch("scripts.ecommerce_report.daily.RuntimeConfig.load")
    def test_failed_day_can_retry_and_success_removes_the_failure_record(
        self, load_config: Mock, run_pipeline_mock: Mock, now: Mock
    ) -> None:
        """Treating a failure file as completion would prevent same-day recovery."""
        load_config.return_value = self.config
        now.return_value = datetime(2026, 8, 11, 9, 7, 5)
        run_pipeline_mock.side_effect = PipelineError(
            "EchoTik采集",
            RuntimeError(
                "cookie=session-secret\n连接超时，请稍后重试\n"
                'File "C:\\private\\runner.py", line 1\nTraceback (most recent call last)'
            ),
        )

        with self.assertRaises(Exception) as raised:
            run_daily_job(self.config_path, self.day)

        self.assertIsInstance(raised.exception, daily_module.DailyJobError)

        failure_path = failure_path_for(self.day, self.config.output_dir)
        self.assertEqual(
            failure_path.read_text(encoding="utf-8"),
            "失败时间: 2026-08-11 09:07:05\n"
            "阶段: EchoTik采集\n"
            "原因: 连接超时，请稍后重试\n",
        )

        output_path = self.config.output_dir / "2026.8.11数据报表.xlsx"

        def successful_retry(config: RuntimeConfig, destination: Path) -> Path:
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_bytes(b"complete")
            return destination

        run_pipeline_mock.side_effect = successful_retry

        result = run_daily_job(self.config_path, self.day)

        self.assertEqual(result, output_path)
        self.assertFalse(failure_path.exists())
        self.assertEqual(run_pipeline_mock.call_count, 2)

    @patch("scripts.ecommerce_report.daily.run_pipeline")
    @patch("scripts.ecommerce_report.daily.RuntimeConfig.load")
    def test_success_creates_only_the_workbook_not_a_success_log(
        self, load_config: Mock, run_pipeline_mock: Mock
    ) -> None:
        """Writing a success marker into the failure directory would misreport status."""
        load_config.return_value = self.config

        def write_report(config: RuntimeConfig, destination: Path) -> Path:
            destination.parent.mkdir(parents=True, exist_ok=True)
            destination.write_bytes(b"complete")
            return destination

        run_pipeline_mock.side_effect = write_report

        result = run_daily_job(self.config_path, self.day)

        self.assertEqual(result.name, "2026.8.11数据报表.xlsx")
        self.assertFalse((self.config.output_dir / "数据报表_失败原因").exists())

    @patch("scripts.ecommerce_report.daily._now")
    def test_configuration_failures_write_to_the_bootstrap_failure_directory(
        self, now: Mock
    ) -> None:
        """Loading config outside the try block would lose early failure records."""
        now.return_value = datetime(2026, 8, 11, 9, 8, 6)
        cases = (
            ("missing.yaml", None, "错误详情已隐藏"),
            ("invalid.yaml", "output_dir: [", None),
            ("invalid-limit.yaml", "detail_limit: 21\n", "detail_limit must be 20"),
        )

        for name, contents, expected_reason in cases:
            with self.subTest(name=name):
                config_path = self.root / name
                if contents is not None:
                    config_path.write_text(contents, encoding="utf-8")

                with self.assertRaises(Exception) as raised:
                    run_daily_job(config_path, self.day)

                expected_path = (
                    self.root
                    / "数据报表_失败原因"
                    / "2026.8.11失败原因.txt"
                )
                self.assertIsInstance(raised.exception, daily_module.DailyJobError)
                self.assertEqual(raised.exception.failure_path, expected_path)
                lines = expected_path.read_text(encoding="utf-8").splitlines()
                self.assertEqual(lines[0], "失败时间: 2026-08-11 09:08:06")
                self.assertEqual(lines[1], "阶段: 读取配置")
                self.assertEqual(len(lines), 3)
                if expected_reason is not None:
                    self.assertEqual(lines[2], f"原因: {expected_reason}")

    def test_skill_local_invalid_config_uses_an_explicit_external_failure_directory(self) -> None:
        """Falling back to config.parent would write a failure artifact into the Skill."""
        skill_config = Path(__file__).resolve().parents[1] / "invalid-config.yaml"
        local_app_data = self.root / "local-app-data"

        with patch.dict(os.environ, {"LOCALAPPDATA": str(local_app_data)}):
            with patch("scripts.ecommerce_report.daily._write_failure") as write_failure:
                with self.assertRaises(daily_module.DailyJobError) as raised:
                    run_daily_job(skill_config, self.day)

        expected = (
            local_app_data
            / "CrossBorderEcommerceDailyReport"
            / "数据报表_失败原因"
            / "2026.8.11失败原因.txt"
        )
        self.assertEqual(raised.exception.failure_path, expected)
        self.assertEqual(write_failure.call_args.args[0], expected)

    def test_sanitize_reason_never_echoes_credentials_or_local_user_paths(self) -> None:
        """Returning any sensitive assignment or local path would leak operator data."""
        secrets = (
            "Authorization: Bearer abc.def.ghi",
            "Bearer abc.def.ghi",
            "Authorization: Basic dXNlcjpwYXNz",
            "https://alice:password@example.com/private",
            "https://alice@example.com/private",
            "password=hunter2",
            "token: token-value",
            "cookie=session-value",
            "secret = private-value",
            "key=private-key",
            r"C:\Users\Alice\private\config.yaml",
            "/home/alice/private/config.yaml",
            "/Users/alice/private/config.yaml",
        )

        for secret in secrets:
            with self.subTest(secret=secret):
                self.assertEqual(
                    daily_module._sanitize_reason(RuntimeError(secret)),
                    "错误详情已隐藏",
                )

        safe_url = "https://example.com/status unavailable"
        self.assertEqual(
            daily_module._sanitize_reason(RuntimeError(safe_url)), safe_url
        )

    def test_sanitize_reason_masks_email_phone_and_url_query_values(self) -> None:
        """Keeping identifiers or query values would leak account data into failure files."""
        reason = daily_module._sanitize_reason(
            RuntimeError(
                "请求 user@example.com 或 13800138000 访问失败: "
                "https://example.com/status?session=private-value&item=42"
            )
        )

        self.assertNotIn("user@example.com", reason)
        self.assertNotIn("13800138000", reason)
        self.assertNotIn("private-value", reason)
        self.assertNotIn("item=42", reason)
        self.assertIn("访问失败", reason)


class EntrypointTests(unittest.TestCase):
    @patch("scripts.run_report.run_pipeline")
    @patch("scripts.run_report.RuntimeConfig.load")
    def test_manual_entrypoint_accepts_config_and_explicit_output(
        self, load_config: Mock, run_pipeline_mock: Mock
    ) -> None:
        """Ignoring --output would write a manual report to the wrong destination."""
        config = Mock(output_dir=Path("default"))
        load_config.return_value = config
        run_pipeline_mock.return_value = Path("chosen.xlsx")

        code = run_report_main(["--config", "settings.yaml", "--output", "chosen.xlsx"])

        self.assertEqual(code, 0)
        load_config.assert_called_once_with(Path("settings.yaml"))
        run_pipeline_mock.assert_called_once_with(config, Path("chosen.xlsx"))

    @patch("scripts.run_report.run_pipeline")
    @patch("scripts.run_report.RuntimeConfig.load")
    def test_manual_entrypoint_uses_the_sanitized_failure_exit(
        self, load_config: Mock, run_pipeline_mock: Mock
    ) -> None:
        """Letting manual failures escape would print traceback, accounts, and paths."""
        load_config.return_value = Mock(output_dir=Path("reports"))
        run_pipeline_mock.side_effect = PipelineError(
            "Amazon采集",
            RuntimeError(
                "user@example.com 13800138000 "
                "https://example.com/status?token=private "
                r"C:\Users\Alice\runner.py"
            ),
        )
        stderr = io.StringIO()

        try:
            with redirect_stderr(stderr):
                code = run_report_main(
                    ["--config", "settings.yaml", "--output", "chosen.xlsx"]
                )
        except Exception as error:
            self.fail(f"manual failure escaped the CLI boundary: {type(error).__name__}")

        text = stderr.getvalue()
        self.assertEqual(code, 1)
        self.assertIn("Amazon采集", text)
        self.assertNotIn("user@example.com", text)
        self.assertNotIn("13800138000", text)
        self.assertNotIn("private", text)
        self.assertNotIn("C:\\Users", text)
        self.assertNotIn("Traceback", text)

    @patch("scripts.run_daily.run_daily_job")
    def test_scheduled_entrypoint_returns_zero_only_on_success(
        self, run_daily_job_mock: Mock
    ) -> None:
        """Returning zero after an exception would conceal a failed scheduled run."""
        run_daily_job_mock.return_value = Path("report.xlsx")
        self.assertEqual(run_daily_main(["--config", "settings.yaml"]), 0)

        run_daily_job_mock.side_effect = RuntimeError("private details")
        stderr = io.StringIO()
        with redirect_stderr(stderr):
            self.assertEqual(run_daily_main(["--config", "settings.yaml"]), 1)
        self.assertNotIn("private details", stderr.getvalue())

    @patch("scripts.run_daily.run_daily_job")
    def test_scheduled_entrypoint_reports_only_a_relative_failure_record_location(
        self, run_daily_job_mock: Mock
    ) -> None:
        """Printing the absolute record path would expose the local account directory."""
        failure_path = Path("C:/runtime/config/数据报表_失败原因/2026.8.11失败原因.txt")
        run_daily_job_mock.side_effect = daily_module.DailyJobError(
            failure_path, "读取配置"
        )
        stderr = io.StringIO()

        with redirect_stderr(stderr):
            code = run_daily_main(["--config", "C:/runtime/config/missing.yaml"])

        self.assertEqual(code, 1)
        self.assertIn("读取配置", stderr.getvalue())
        self.assertIn("2026.8.11失败原因.txt", stderr.getvalue())
        self.assertNotIn("C:/runtime", stderr.getvalue())


class SchedulerScriptTests(unittest.TestCase):
    def setUp(self) -> None:
        self.script_path = (
            Path(__file__).resolve().parents[1]
            / "scripts"
            / "install_scheduled_task.ps1"
        )

    def test_windows_powershell_parsefile_accepts_the_ascii_script(self) -> None:
        """Adding UTF-8 text without a BOM would break Windows PowerShell 5.1 ParseFile."""
        self.script_path.read_bytes().decode("ascii")
        command = (
            "$tokens=$null;$errors=$null;"
            "[System.Management.Automation.Language.Parser]::ParseFile("
            f"'{self.script_path}',[ref]$tokens,[ref]$errors)>$null;"
            "if($errors.Count){$errors|% Message;exit 1}"
        )

        result = subprocess.run(
            ["powershell", "-NoProfile", "-Command", command],
            capture_output=True,
            text=True,
            check=False,
        )

        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)

    def test_scheduler_uses_a_preflighted_absolute_python_executable(self) -> None:
        """Using PATH's first python directly could register a broken task action."""
        source = self.script_path.read_text(encoding="ascii")

        self.assertIn("[string] $PythonExecutable", source)
        self.assertIn("$ResolvedPythonExecutable", source)
        self.assertIn("Resolve-Path -LiteralPath $PythonExecutable", source)
        self.assertIn("import yaml", source)
        self.assertIn("import playwright", source)
        self.assertIn("spec_from_file_location", source)
        self.assertIn("$LASTEXITCODE", source)
        self.assertIn(
            "New-ScheduledTaskAction -Execute $ResolvedPythonExecutable", source
        )
        self.assertLess(
            source.index("spec_from_file_location"),
            source.index("New-ScheduledTaskAction"),
        )
        self.assertLess(
            source.index("Write-Host \"Proposed task action:"),
            source.index("Register-ScheduledTask"),
        )


if __name__ == "__main__":
    unittest.main()
